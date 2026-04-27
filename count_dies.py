#!/usr/bin/env python3
"""
统计111.xlsx中的刀模在所有生产工艺单中出现的次数
按年份分组统计（JF23=23年, JF24=24年, JF25=25年, JF26=26年）
一个计划跟踪号算1次
"""

import openpyxl
import re
from pathlib import Path

BASE_DIR = Path('/home/smg/claude/time')
DIE_FILE = BASE_DIR / '111.xlsx'
PRODUCTION_FILES = sorted(BASE_DIR.glob('生产工艺单JF*.xlsx'))
OUTPUT_FILE = BASE_DIR / '刀模使用次数统计.xlsx'

# 列索引 (0-based)
COL_PLAN_TRACK = 0   # 计划跟踪号
COL_QIAN = 4         # 前片
COL_BIAN = 5         # 边片
COL_HOU = 6          # 后片


def normalize_code(code):
    """标准化：去除#、cm（不区分大小写）、空格，转小写"""
    if not code:
        return ''
    code = code.replace('#', '')
    code = re.sub(r'cm', '', code, flags=re.IGNORECASE)
    code = code.replace(' ', '')
    return code.lower()


def extract_code(s):
    """提取中文之前的代码"""
    if not s:
        return None
    match = re.match(r'^([^\u4e00-\u9fff]+)', str(s))
    return match.group(1) if match else None


def get_year(file_path):
    """从文件名提取年份"""
    name = file_path.name
    if 'JF23' in name:
        return '23年'
    elif 'JF24' in name:
        return '24年'
    elif 'JF25' in name:
        return '25年'
    elif 'JF26' in name:
        return '26年'
    return name


def main():
    # 读取111.xlsx刀模列表（所有行，包含重复）
    print("读取刀模列表...")
    wb = openpyxl.load_workbook(DIE_FILE, read_only=True, data_only=True)
    ws = wb.active
    target_dies = []
    for row in ws.iter_rows(values_only=True):
        if row[0]:
            target_dies.append(str(row[0]).strip())
    wb.close()

    print(f"共 {len(target_dies)} 个刀模（含重复）")
    print(f"唯一刀模数: {len(set(target_dies))}")

    # 创建标准化的刀模列表
    target_normalized_list = [(normalize_code(die), die) for die in target_dies]

    # 统计结果 {刀模: {'23年': set(), '24年': set(), ...}}
    years = ['23年', '24年', '25年', '26年']
    results = {die: {y: set() for y in years} for die in target_dies}

    # 遍历所有生产工艺单
    for file_path in PRODUCTION_FILES:
        year = get_year(file_path)
        print(f"处理: {file_path.name} ({year})...")
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            plan_track = row[COL_PLAN_TRACK]
            if not plan_track:
                continue

            # 提取前片、边片、后片的刀模代码
            codes_in_row = set()
            for col in [COL_QIAN, COL_BIAN, COL_HOU]:
                value = row[col] if len(row) > col else None
                code = extract_code(value)
                if code:
                    code_norm = normalize_code(code)
                    for norm, orig_die in target_normalized_list:
                        if norm == code_norm:
                            codes_in_row.add(orig_die)

            # 记录这个计划跟踪号匹配到的刀模
            for orig_die in codes_in_row:
                results[orig_die][year].add(plan_track)

        wb.close()

    # 生成结果文件
    print(f"\n生成结果文件: {OUTPUT_FILE}...")
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "刀模统计"

    # 写入表头
    ws_out.append(["刀模号", "23年", "24年", "25年", "26年", "合计"])

    # 写入数据
    total_all = 0
    dies_found = 0
    for die in target_dies:
        counts = {y: len(results[die][y]) for y in years}
        total = sum(counts.values())
        total_all += total
        if total > 0:
            dies_found += 1
        ws_out.append([die, counts['23年'], counts['24年'], counts['25年'], counts['26年'], total])

    wb_out.save(OUTPUT_FILE)

    print(f"\n完成！")
    print(f"111.xlsx刀模总数: {len(target_dies)}")
    print(f"唯一刀模数: {len(set(target_dies))}")
    print(f"匹配到: {dies_found}")
    print(f"未匹配: {len(target_dies) - dies_found}")
    print(f"总计: {total_all} 次出现")
    print(f"结果已保存到: {OUTPUT_FILE}")


if __name__ == '__main__':
    main()
