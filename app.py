"""
刀模管理系统 Web 应用
"""
import sys
import re
import glob
from pathlib import Path
from collections import defaultdict
from datetime import datetime, date, timedelta

from flask import Flask, render_template, request, jsonify, send_file, session, redirect, url_for, flash
from functools import wraps
import openpyxl

# 复用匹配逻辑
from match_die_molds import (
    normalize_process_ref, get_die_mold_keys, count_model_segments, is_v_type,
    is_v_used, is_nonv_used, load_die_molds as ml_load_die_molds,
    load_process_refs as ml_load_process_refs, match_die_molds as ml_match_die_molds
)
from config import BASE_DIR, DIE_MOLD_FILE, DIE_MOLD_SHEET
from utils.excel_reader import (
    load_die_molds, add_die_mold, update_die_mold,
    delete_die_mold, search_die_molds, offline_die_mold
)
from utils.excel_writer import export_to_excel, create_match_report
from utils.borrow_records import (
    init_borrow_db, add_borrow_record, return_mold,
    get_borrow_records, get_borrow_record_by_id,
    get_borrow_summary_by_model, get_borrow_summary_by_model_position,
    delete_borrow_record
)
from utils.process_cache import (
    search_by_track_no_cached, rebuild_cache, get_cache_stats, init_cache_db
)
from utils.auth import init_user_db, verify_user, create_user, get_all_users, delete_user, change_password, change_password_by_admin

# 初始化用户数据库
init_user_db()

# 登录验证装饰器
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

# 预编译正则表达式：提取刀模前缀（如 A2-523, A1K1-C-57）
_PREFIX_RE = re.compile(r'^([A-Za-z0-9]+-[A-Za-z0-9]+)')
# 去掉括号及括号内容，如 (改版)、(还好用)、(改) 等，括号内是说明不是另一个刀模
_REVISION_RE = re.compile(r"\s*[\(（][^\)）]*[\)）]\s*")

app = Flask(__name__)
app.config.from_object('config')


# ── 首页/搜索 ───────────────────────────────────────────
@app.route('/')
@login_required
def index():
    return render_template('index.html')


# ── 登录/登出 ───────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        if verify_user(username, password):
            session['user_id'] = username
            session['username'] = username
            return redirect(url_for('index'))
        flash('用户名或密码错误')
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


# ── 用户管理 ───────────────────────────────────────────
@app.route('/users')
@login_required
def users_page():
    """用户管理页面"""
    return render_template('users.html')


@app.route('/api/users', methods=['GET'])
@login_required
def api_get_users():
    """获取用户列表"""
    users = get_all_users()
    return jsonify({"success": True, "users": users})


@app.route('/api/users', methods=['POST'])
@login_required
def api_create_user():
    """创建新用户"""
    data = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')

    success, message = create_user(username, password)
    if success:
        return jsonify({"success": True, "message": message})
    return jsonify({"success": False, "error": message}), 400


@app.route('/api/users/<username>/password', methods=['PUT'])
@login_required
def api_change_password(username):
    """修改用户密码"""
    data = request.get_json()
    old_password = data.get('old_password', '')
    new_password = data.get('new_password', '')

    # 检查权限：只能修改自己的密码，或者管理员可以修改任意用户密码
    current_user = session.get('username', '')
    if current_user != username:
        return jsonify({"success": False, "error": "无权修改其他用户的密码"}), 403

    success, message = change_password(username, old_password, new_password)
    if success:
        return jsonify({"success": True, "message": message})
    return jsonify({"success": False, "error": message}), 400


@app.route('/api/users/<username>', methods=['DELETE'])
@login_required
def api_delete_user(username):
    """删除用户"""
    success, message = delete_user(username)
    if success:
        return jsonify({"success": True, "message": message})
    return jsonify({"success": False, "error": message}), 400


@app.route('/inventory')
@login_required
def inventory_page():
    return render_template('inventory.html')


@app.route('/upload')
@login_required
def upload_page():
    return render_template('upload.html')


@app.route('/borrow')
@login_required
def borrow_page():
    # 初始化借出数据库
    init_borrow_db()
    return render_template('borrow.html')


@app.route('/report')
@login_required
def report_page():
    return render_template('report.html')


# ── API: 计划跟踪号搜索 ─────────────────────────────────
# 全局缓存：预加载刀模库存
_inventory_cache = {
    'records': None,
    'models_index': None,
    'v_models': None,
    'nonv_prefix_index': None,  # 前缀分组索引
    'nonv_full_index': None,    # 完整键索引（用于后备）
}

# 工艺单文件缓存（按年份）
_process_files_cache = {
    'files': None,
    'year_index': {},  # {'JF26': [Path('生产工艺单JF26.xlsx'), ...]}
}

def _extract_model_prefix(model):
    """提取刀模前缀（如 A2-523-58-037 → A2-523）"""
    m = _PREFIX_RE.match(model)
    return m.group(1) if m else None


def _get_process_files_by_year():
    """获取按年份索引的工艺单文件"""
    if _process_files_cache['files'] is None:
        files = sorted(BASE_DIR.glob("生产工艺单*.xlsx"))
        _process_files_cache['files'] = files
        
        # 构建年份索引
        year_index = defaultdict(list)
        for f in files:
            # 从文件名提取年份，如 生产工艺单JF26.xlsx → JF26
            m = re.match(r'生产工艺单(JF\d+)\.xlsx', f.name)
            if m:
                year_key = m.group(1)
                year_index[year_key].append(f)
        
        _process_files_cache['year_index'] = dict(year_index)
    
    return _process_files_cache['year_index'], _process_files_cache['files']


def _search_track_no_in_file(fpath, track_no):
    """在指定工艺单文件中搜索计划跟踪号"""
    wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
    if "Sheet1" not in wb.sheetnames:
        wb.close()
        return None
    ws = wb["Sheet1"]

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and str(row[0]).strip() == track_no:
            result = {
                "track_no": str(row[0]).strip(),
                "customer": str(row[1]).strip() if len(row) > 1 and row[1] else None,
                "helmet_type": str(row[7]).strip() if len(row) > 7 and row[7] else None,
                "cap_type": str(row[8]).strip() if len(row) > 8 and row[8] else None,
                "eyebrow_top": str(row[2]).strip() if len(row) > 2 and row[2] else None,
                "eyebrow_bot": str(row[3]).strip() if len(row) > 3 and row[3] else None,
                "front": str(row[4]).strip() if len(row) > 4 and row[4] else None,
                "side": str(row[5]).strip() if len(row) > 5 and row[5] else None,
                "back": str(row[6]).strip() if len(row) > 6 and row[6] else None,
            }
            wb.close()
            return result
    
    wb.close()
    return None


def _search_track_no_optimized(track_no):
    """优化的计划跟踪号搜索：按年份优先查找"""
    # 从计划跟踪号提取年份前缀，如 JF2603520 → JF26
    year_match = re.match(r'^(JF\d{2})', track_no.strip())
    
    if year_match:
        year_key = year_match.group(1)
        year_index, all_files = _get_process_files_by_year()
        
        # 1. 优先在对应年份文件中查找
        if year_key in year_index:
            for fpath in year_index[year_key]:
                result = _search_track_no_in_file(fpath, track_no)
                if result:
                    return result
        
        # 2. 如果没找到，在其他年份文件中查找
        for key, files in year_index.items():
            if key != year_key:
                for fpath in files:
                    result = _search_track_no_in_file(fpath, track_no)
                    if result:
                        return result
    else:
        # 没有年份前缀，遍历所有文件
        _, all_files = _get_process_files_by_year()
        for fpath in all_files:
            result = _search_track_no_in_file(fpath, track_no)
            if result:
                return result
    
    return None

def get_inventory_cache():
    """获取预加载的刀模库存缓存（优化版：前缀分组索引）"""
    if _inventory_cache['records'] is None:
        inventory = ml_load_die_molds(DIE_MOLD_FILE, DIE_MOLD_SHEET)

        # 加载借出状态
        try:
            borrow_summary = get_borrow_summary_by_model()
        except:
            borrow_summary = {}

        # 给每条库存记录添加借出信息
        for rec in inventory:
            mold_model = rec.get('model', '')
            borrow_info = borrow_summary.get(mold_model)
            if borrow_info:
                rec['is_borrowed'] = True
                rec['borrower'] = borrow_info.get('borrower', '')
                rec['borrow_date'] = borrow_info.get('borrow_date', '')
            else:
                rec['is_borrowed'] = False
                rec['borrower'] = ''
                rec['borrow_date'] = ''

        # 构建索引
        inventory_models = {rec['model']: rec for rec in inventory}

        # V型刀模列表
        v_models = [rec for rec in inventory if rec['model'].startswith('V')]

        # 非V型刀模：前缀分组索引
        nonv_prefix_index = defaultdict(list)
        for rec in inventory:
            if not rec['model'].startswith('V'):
                prefix = _extract_model_prefix(rec['model'])
                if prefix:
                    nonv_prefix_index[prefix].append(rec)

        # 非V型刀模：完整键索引（后备方案）
        nonv_full_index = {}
        for rec in inventory:
            if not rec['model'].startswith('V'):
                keys = get_die_mold_keys(rec['model'])
                for key in keys:
                    if key not in nonv_full_index:
                        nonv_full_index[key] = rec

        _inventory_cache['records'] = inventory
        _inventory_cache['models_index'] = inventory_models
        _inventory_cache['v_models'] = v_models
        _inventory_cache['nonv_prefix_index'] = dict(nonv_prefix_index)
        _inventory_cache['nonv_full_index'] = nonv_full_index

    return _inventory_cache


def invalidate_inventory_cache():
    """清除刀模库存缓存（在增删改后调用）"""
    _inventory_cache['records'] = None
    _inventory_cache['models_index'] = None
    _inventory_cache['v_models'] = None
    _inventory_cache['nonv_prefix_index'] = None
    _inventory_cache['nonv_full_index'] = None


@app.route('/api/process/search')
def api_search_process():
    """根据计划跟踪号搜索工艺单（优化版：按年份优先查找 + 始终返回刀模信息）"""
    track_no = request.args.get('track_no', '').strip()
    if not track_no:
        return jsonify({"error": "计划跟踪号不能为空"}), 400

    # 优先从缓存查询
    cached = search_by_track_no_cached(track_no, auto_update=True)
    if cached:
        # 使用预加载的库存缓存
        cache = get_inventory_cache()
        inventory_models = cache['models_index']
        v_models = cache['v_models']
        nonv_prefix_index = cache['nonv_prefix_index']
        nonv_full_index = cache['nonv_full_index']

        result = {
            "track_no": cached["track_no"],
            "customer": cached["customer"],
            "helmet_type": cached["helmet_type"],
            "cap_type": cached["cap_type"],
            "molds": []
        }

        # 处理上眉
        if cached["eyebrow_top"]:
            mold_infos = check_mold_in_stock_fast(
                cached["eyebrow_top"], "上眉", "V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        # 处理下眉
        if cached["eyebrow_bot"]:
            mold_infos = check_mold_in_stock_fast(
                cached["eyebrow_bot"], "下眉", "V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        # 处理前片
        if cached["front"]:
            mold_infos = check_mold_in_stock_fast(
                cached["front"], "前片", "非V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        # 处理边片
        if cached["side"]:
            mold_infos = check_mold_in_stock_fast(
                cached["side"], "边片", "非V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        # 处理后片
        if cached["back"]:
            mold_infos = check_mold_in_stock_fast(
                cached["back"], "后片", "非V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        return jsonify(result)

    # 缓存未命中，使用优化的文件搜索（按年份优先）
    cached = _search_track_no_optimized(track_no)
    if cached:
        # 使用预加载的库存缓存
        cache = get_inventory_cache()
        inventory_models = cache['models_index']
        v_models = cache['v_models']
        nonv_prefix_index = cache['nonv_prefix_index']
        nonv_full_index = cache['nonv_full_index']

        result = {
            "track_no": cached["track_no"],
            "customer": cached["customer"],
            "helmet_type": cached["helmet_type"],
            "cap_type": cached["cap_type"],
            "molds": []
        }

        # 处理上眉
        if cached["eyebrow_top"]:
            mold_infos = check_mold_in_stock_fast(
                cached["eyebrow_top"], "上眉", "V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        # 处理下眉
        if cached["eyebrow_bot"]:
            mold_infos = check_mold_in_stock_fast(
                cached["eyebrow_bot"], "下眉", "V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        # 处理前片
        if cached["front"]:
            mold_infos = check_mold_in_stock_fast(
                cached["front"], "前片", "非V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        # 处理边片
        if cached["side"]:
            mold_infos = check_mold_in_stock_fast(
                cached["side"], "边片", "非V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        # 处理后片
        if cached["back"]:
            mold_infos = check_mold_in_stock_fast(
                cached["back"], "后片", "非V型",
                inventory_models, v_models, nonv_prefix_index, nonv_full_index
            )
            result["molds"].extend(mold_infos)

        return jsonify(result)

    # 未找到
    return jsonify({"error": "未找到该计划跟踪号"}), 404


def _search_process_slow(track_no):
    """直接搜索工艺单文件（慢，作为回退）"""
    files = sorted(BASE_DIR.glob("生产工艺单*.xlsx"))
    if not files:
        return jsonify({"error": "未找到工艺单文件"}), 404

    for fpath in files:
        wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
        if "Sheet1" not in wb.sheetnames:
            wb.close()
            continue
        ws = wb["Sheet1"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and str(row[0]).strip() == track_no:
                eyebrow_top = row[2] if len(row) > 2 else None
                eyebrow_bot = row[3] if len(row) > 3 else None
                front = row[4] if len(row) > 4 else None
                side = row[5] if len(row) > 5 else None
                back = row[6] if len(row) > 6 else None

                result = {
                    "track_no": row[0],
                    "customer": row[1] if len(row) > 1 else None,
                    "helmet_type": row[7] if len(row) > 7 else None,
                    "cap_type": row[8] if len(row) > 8 else None,
                    "molds": []
                }

                inventory = ml_load_die_molds(DIE_MOLD_FILE, DIE_MOLD_SHEET)
                inventory_models = {rec['model']: rec for rec in inventory}

                if eyebrow_top and str(eyebrow_top).strip():
                    result["molds"].append(check_mold_in_stock(str(eyebrow_top).strip(), "上眉", "V型", inventory_models))
                if eyebrow_bot and str(eyebrow_bot).strip():
                    result["molds"].append(check_mold_in_stock(str(eyebrow_bot).strip(), "下眉", "V型", inventory_models))
                if front and str(front).strip():
                    result["molds"].append(check_mold_in_stock(str(front).strip(), "前片", "非V型", inventory_models))
                if side and str(side).strip():
                    result["molds"].append(check_mold_in_stock(str(side).strip(), "边片", "非V型", inventory_models))
                if back and str(back).strip():
                    result["molds"].append(check_mold_in_stock(str(back).strip(), "后片", "非V型", inventory_models))

                wb.close()
                return jsonify(result)

        wb.close()

    return jsonify({"error": "未找到该计划跟踪号"}), 404


@app.route('/api/process/cache/rebuild', methods=['POST'])
def api_rebuild_cache():
    """重建工艺单缓存"""
    try:
        count = rebuild_cache()
        return jsonify({"success": True, "count": count})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/process/cache/status')
def api_cache_status():
    """获取缓存状态"""
    stats = get_cache_stats()
    return jsonify(stats)


@app.route('/api/inventory/cache/invalidate', methods=['POST'])
def api_inventory_cache_invalidate():
    """清除刀模库存缓存"""
    invalidate_inventory_cache()
    return jsonify({"success": True, "message": "库存缓存已清除"})


@app.route('/api/refresh/all', methods=['POST'])
def api_refresh_all():
    """刷新系统数据：同时清除库存缓存和重建工艺单缓存

    用于手动更新刀模汇总表、生产工艺单、大货生产计划跟踪表后，
    点击刷新按钮使系统数据立即更新。
    """
    try:
        # 1. 清除库存缓存
        invalidate_inventory_cache()

        # 2. 重建工艺单缓存
        cache_count = rebuild_cache()

        return jsonify({
            "success": True,
            "message": "系统数据已刷新",
            "details": {
                "inventory_cache": "已清除",
                "process_cache": f"已重建，共 {cache_count:,} 条记录"
            }
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/process/batch_upload', methods=['POST'])
def api_batch_upload():
    """批量上传计划跟踪号Excel文件，返回计划跟踪号和领用人列表"""
    if 'file' not in request.files:
        return jsonify({"error": "没有上传文件"}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "只支持 .xlsx 或 .xls 文件"}), 400

    try:
        # 保存临时文件
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        temp_file = BASE_DIR / f"temp_batch_{timestamp}.xlsx"
        file.save(temp_file)

        # 读取第一列的计划跟踪号和第二列的领用人
        wb = openpyxl.load_workbook(temp_file, data_only=True, read_only=True)
        ws = wb.active

        track_nos = []
        borrower_map = {}  # {计划跟踪号: 领用人}
        for row in ws.iter_rows(min_row=1, values_only=True):
            if row and row[0]:
                track_no = str(row[0]).strip()
                if track_no:
                    track_nos.append(track_no)
                    borrower = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                    borrower_map[track_no] = borrower

        wb.close()
        temp_file.unlink()  # 删除临时文件

        return jsonify({"track_nos": track_nos, "borrower_map": borrower_map})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/process/batch_export', methods=['POST'])
def api_batch_export():
    """批量导出查询结果到Excel - 生成两个Sheet
    Sheet1: 标准格式（原始数据，不合并）
    Sheet2: 按领用人分组，位置和备注合并显示
    """
    data = request.get_json()
    results = data.get('results', [])

    if not results:
        return jsonify({"error": "没有可导出的数据"}), 400

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from collections import defaultdict

    # ── 第一步：合并 Sheet2 数据 ──────────────────────────────
    merged_results = defaultdict(lambda: {
        'track_no': '', 'process_model': '', 'type_name': '', 'positions': [],
        'inventory_model': '', 'notes': [], 'in_stock': False, 'borrower': ''
    })

    for r in results:
        key = (r.get('track_no', ''), r.get('process_model', ''), r.get('type_name', ''))
        entry = merged_results[key]
        entry['track_no'] = r.get('track_no', '')
        entry['process_model'] = r.get('process_model', '')
        entry['type_name'] = r.get('type_name', '')
        entry['borrower'] = r.get('borrower', '')
        entry['inventory_model'] = r.get('inventory_model', '') or entry['inventory_model']

        pos = r.get('position', '').strip()
        if pos:
            entry['positions'].append(pos)

        note = r.get('note', '').strip()
        if note:
            entry['notes'].append(note)

        if r.get('in_stock'):
            entry['in_stock'] = True

    # 合并后的列表
    merged_list = []
    for key, entry in merged_results.items():
        merged_list.append({
            'track_no': entry['track_no'],
            'process_model': entry['process_model'],
            'type_name': entry['type_name'],
            'positions': entry['positions'],
            'position': ' | '.join(entry['positions']),
            'notes': entry['notes'],
            'note': ' | '.join(entry['notes']),
            'inventory_model': entry['inventory_model'],
            'in_stock': entry['in_stock'],
            'borrower': entry['borrower']
        })

    # ── 第二步：计算预警信息 ──────────────────────────────────
    def calc_total_quantity(positions_list, notes_list):
        """计算刀模总套数 - 每一行至少1套，备注* N表示该行有N套"""
        if not positions_list:
            return 0
        total = 0
        for i, pos in enumerate(positions_list):
            total += 1  # 每行至少1套
            note = notes_list[i] if i < len(notes_list) else ''
            match = re.search(r'\*(\d+)', str(note))
            if match:
                total += int(match.group(1)) - 1  # 加额外套数
        return total

    # 统计使用人数：按(工艺单型号, 领用人)去重，同一人同一工艺单只算1人
    usage_count_map = defaultdict(set)  # {process_model: {borrowers...}}
    for item in merged_list:
        if item['process_model'] and item['borrower']:
            usage_count_map[item['process_model']].add(item['borrower'])

    # 为每个 merged_list 条目添加预警信息
    for item in merged_list:
        total_qty = calc_total_quantity(item['positions'], item['notes'])
        borrowers = usage_count_map.get(item['process_model'], set())
        usage_count = len(borrowers)

        item['total_quantity'] = total_qty
        item['usage_count'] = usage_count

        # 预警条件：使用人数 > 刀模总套数
        if item['in_stock'] and usage_count > total_qty:
            borrowers_str = '、'.join(sorted(borrowers))
            item['warning'] = f'⚠️预警:{borrowers_str}需求{usage_count}套>库存{total_qty}套'
        else:
            item['warning'] = ''

    # ── 样式定义 ─────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    even_row_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    odd_row_fill = PatternFill(fill_type="solid", fgColor="FFFFFF")
    missing_row_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")

    status_available_font = Font(color="006100", bold=True, size=10)
    status_missing_font = Font(color="9C0006", bold=True, size=10)

    thin_border = Border(
        left=Side(style='thin', color='B8B8B8'),
        right=Side(style='thin', color='B8B8B8'),
        top=Side(style='thin', color='B8B8B8'),
        bottom=Side(style='thin', color='B8B8B8')
    )
    header_border = Border(
        left=Side(style='medium', color='FFFFFF'),
        right=Side(style='medium', color='FFFFFF'),
        top=Side(style='medium', color='FFFFFF'),
        bottom=Side(style='medium', color='FFFFFF')
    )

    # ── 创建Excel ───────────────────────────────────────────
    wb = openpyxl.Workbook()

    # ═══════════════════════════════════════════════════════════
    # Sheet1: 标准格式（位置已合并）
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "刀模批量查询结果"

    headers = ["计划跟踪号", "工艺单型号", "部位", "位置", "库存型号", "原备注", "状态"]
    ws1.append(headers)
    ws1.row_dimensions[1].height = 25

    for col_idx, header in enumerate(headers, start=1):
        cell = ws1.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    in_stock_count = 0
    missing_count = 0

    for row_idx, row_data in enumerate(results, start=2):
        row = [
            row_data.get('track_no', ''),
            row_data.get('process_model', ''),
            row_data.get('type_name', ''),
            row_data.get('position', ''),
            row_data.get('inventory_model', ''),
            row_data.get('note', ''),
            '有刀模' if row_data.get('in_stock') else '缺失'
        ]
        ws1.append(row)

        in_stock = row_data.get('in_stock', False)
        if in_stock:
            in_stock_count += 1
        else:
            missing_count += 1

        ws1.row_dimensions[row_idx].height = 20

        if not in_stock:
            row_fill = missing_row_fill
            status_font = status_missing_font
        else:
            row_fill = even_row_fill if row_idx % 2 == 0 else odd_row_fill
            status_font = status_available_font

        for col_idx in range(1, len(headers) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            cell.fill = row_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        ws1.cell(row=row_idx, column=len(headers)).font = status_font
        ws1.cell(row=row_idx, column=4).alignment = Alignment(horizontal='left', vertical='center')
        ws1.cell(row=row_idx, column=5).alignment = Alignment(horizontal='left', vertical='center')
        ws1.cell(row=row_idx, column=6).alignment = Alignment(horizontal='left', vertical='center')

    # 统计行
    summary_row = len(results) + 2
    ws1.row_dimensions[summary_row].height = 22
    ws1.append([f"共 {len(merged_list)} 条记录", f"有刀模: {in_stock_count} 条", f"缺失: {missing_count} 条", "", "", "", ""])

    for col_idx in range(1, len(headers) + 1):
        cell = ws1.cell(row=summary_row, column=col_idx)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws1.merge_cells(f'A{summary_row}:B{summary_row}')
    ws1.merge_cells(f'C{summary_row}:D{summary_row}')
    ws1.merge_cells(f'E{summary_row}:G{summary_row}')

    ws1.column_dimensions["A"].width = 18
    ws1.column_dimensions["B"].width = 24
    ws1.column_dimensions["C"].width = 8
    ws1.column_dimensions["D"].width = 20
    ws1.column_dimensions["E"].width = 24
    ws1.column_dimensions["F"].width = 20
    ws1.column_dimensions["G"].width = 10
    ws1.freeze_panes = 'A2'

    # ═══════════════════════════════════════════════════════════
    # Sheet2: 按领用人分组（可折叠）
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="按领用人分组")

    borrower_groups = defaultdict(list)
    for r in merged_list:
        borrower = r.get('borrower', '') or '未指定'
        borrower_groups[borrower].append(r)

    title_font = Font(bold=True, color="FFFFFF", size=12)
    title_fill = PatternFill(fill_type="solid", fgColor="2F5496")

    ws2_headers = ["领用人", "计划跟踪号", "工艺单型号", "部位", "库存型号", "位置", "原备注", "状态", "预警"]
    ws2.append(ws2_headers)
    ws2.row_dimensions[1].height = 25

    for col_idx, header in enumerate(ws2_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border

    # 开启行折叠功能
    ws2.sheet_properties.outlinePr.summaryBelow = False

    current_row = 2
    group_ranges = []  # [(组起始行, 组结束行), ...]

    for borrower in sorted(borrower_groups.keys()):
        items = borrower_groups[borrower]

        # 领用人标题行
        ws2.row_dimensions[current_row].height = 22
        ws2.merge_cells(f'A{current_row}:I{current_row}')
        title_cell = ws2.cell(row=current_row, column=1, value=f"领用人: {borrower} (共 {len(items)} 条刀模)")
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        for c in range(1, 10):
            ws2.cell(row=current_row, column=c).border = thin_border
        current_row += 1

        # 按计划跟踪号排序（计划跟踪号优先，部位次之）
        part_order = {'上眉': 0, '下眉': 1, '前片': 2, '边片': 3, '后片': 4}
        sorted_items = sorted(items, key=lambda x: (x.get('track_no', ''), part_order.get(x.get('type_name', ''), 99)))

        group_start = current_row
        for idx, item in enumerate(sorted_items):
            ws2.row_dimensions[current_row].height = 18
            in_stock = item.get('in_stock', False)

            if in_stock:
                row_fill = even_row_fill if idx % 2 == 0 else odd_row_fill
                status_text = '有刀模'
                status_font = status_available_font
            else:
                row_fill = missing_row_fill
                status_text = '缺失'
                status_font = status_missing_font

            warning_text = item.get('warning', '')
            warning_fill = PatternFill(fill_type="solid", fgColor="FF6B6B") if warning_text else row_fill
            warning_font = Font(color="9C0006", bold=True, size=10) if warning_text else None

            row_data = [
                '',
                item.get('track_no', ''),
                item.get('process_model', ''),
                item.get('type_name', ''),
                item.get('inventory_model', ''),
                item.get('position', ''),
                item.get('note', ''),
                status_text,
                warning_text
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws2.cell(row=current_row, column=col_idx, value=value)
                cell.fill = row_fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if col_idx == 2 or col_idx == 4 or col_idx == 8:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                if col_idx == 8:
                    cell.font = status_font
                if col_idx == 9 and warning_text:
                    cell.fill = warning_fill
                    cell.font = warning_font

            current_row += 1

        group_end = current_row - 1
        group_ranges.append((group_start, group_end))
        current_row += 1  # 空行分隔

    # 创建折叠组（从后往前添加，避开空行）
    for start, end in group_ranges:
        ws2.row_dimensions[start].outlineLevel = 1
        ws2.row_dimensions[start].collapsed = True
        for r in range(start + 1, end + 1):
            ws2.row_dimensions[r].outlineLevel = 1

    ws2.column_dimensions["A"].width = 12
    ws2.column_dimensions["B"].width = 18
    ws2.column_dimensions["C"].width = 22
    ws2.column_dimensions["D"].width = 8
    ws2.column_dimensions["E"].width = 22
    ws2.column_dimensions["F"].width = 20
    ws2.column_dimensions["G"].width = 20
    ws2.column_dimensions["H"].width = 10
    ws2.column_dimensions["I"].width = 26
    ws2.freeze_panes = 'A2'

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = BASE_DIR / f"刀模批量查询结果_{timestamp}.xlsx"
    wb.save(output_path)

    return send_file(output_path, as_attachment=True)


def check_mold_in_stock_fast(model, type_name, mold_type, inventory_models, v_models, nonv_prefix_index, nonv_full_index):
    """快速检查刀模是否在库存中（优化版：前缀分组索引）

    返回所有匹配的记录列表，支持同一刀模在多个位置的情况。

    非V型刀模优化策略：
    1. 从工艺单型号提取前缀（如 A2-523#-57CM前片 → A2-523）
    2. 只在前缀分组内查找（如 A2-523 分组平均只有5个刀模）
    3. 如果前缀分组没找到，才使用完整索引作为后备
    4. 非V型刀模必须按前/边/后关键字匹配：边片只能匹配备注含"边"或"侧"的刀模
    """
    # 标准化处理
    normalized = normalize_process_ref(model) if mold_type == "非V型" else model.strip()

    found_recs = []

    if mold_type == "V型":
        # V型刀模匹配：使用预加载的V型列表
        found_recs = _match_v_mold_fast(model.strip(), v_models)
    else:
        # 非V型刀模匹配：优先使用前缀分组索引（需要传入type_name用于关键字匹配）
        found_recs = _match_nonv_mold_by_prefix(normalized, type_name, nonv_prefix_index, nonv_full_index)

    if found_recs:
        # 返回所有匹配的记录
        results = []
        for found_rec in found_recs:
            results.append({
                "model": model,
                "inventory_model": found_rec.get("model"),
                "type_name": type_name,
                "mold_type": mold_type,
                "in_stock": True,
                "position": found_rec.get("position"),
                "quantity": found_rec.get("available", found_rec.get("quantity")),
                "total_quantity": found_rec.get("total_quantity", found_rec.get("quantity")),
                "borrowed": found_rec.get("borrowed", 0),
                "available": found_rec.get("available", found_rec.get("quantity")),
                "note": found_rec.get("note"),
                "borrower": found_rec.get("borrower"),
                "is_borrowed": found_rec.get("is_borrowed", False),
                "borrow_date": found_rec.get("borrow_date", ""),
            })
        return results
    else:
        return [{
            "model": model,
            "inventory_model": None,
            "type_name": type_name,
            "mold_type": mold_type,
            "in_stock": False,
            "position": None,
            "quantity": None,
            "total_quantity": None,
            "borrowed": None,
            "available": None,
            "note": None,
            "borrower": None,
            "is_borrowed": False,
            "borrow_date": "",
        }]


def _match_nonv_mold_by_prefix(normalized_model, type_name, nonv_prefix_index, nonv_full_index):
    """非V型刀模匹配：优先使用前缀分组索引（优化版）

    返回所有匹配的记录列表。

    策略：
    1. 从标准化型号提取前缀（如 A2-523-57 → A2-523）
    2. 只在该前缀分组内查找（通常只有2-10个刀模）
    3. 如果前缀分组没找到，使用完整索引作为后备
    4. 段数必须匹配：一段只能匹配一段，二段只能匹配二段，三段只能匹配三段
    5. 关键字必须匹配：前片只能匹配备注含"前"的刀模，边片只能匹配备注含"边"或"侧"的刀模，后片只能匹配备注含"后"的刀模
    """
    # 提取关键字
    keyword_required = None
    if "前" in type_name:
        keyword_required = "前"
    elif "边" in type_name:
        keyword_required = "边"  # "边" 或 "侧" 都可以
    elif "后" in type_name:
        keyword_required = "后"

    def note_has_keyword(rec, required):
        """检查备注或刀模型号是否包含关键字（边包含侧）

        当刀模型号本身包含关键字时（如 A4-084-58.5双针前），也视为匹配。
        这符合用户的使用习惯：刀模型号写 A4-084-58.5双针，备注写 前。

        特殊兜底：如果刀型号号中没有前/边/后关键字、数量>=3、备注为空或备注中没有前/边/后，
        则视为通用刀模，可以匹配任何类型的前/边/后片需求。
        """
        if not required:
            return True
        note = rec.get('note', '') or ''
        model = rec.get('model', '') or ''
        note_upper = note.upper()
        model_upper = model.upper()

        # 如果备注或型号包含关键字，直接匹配
        if required == "边":
            if "边" in note_upper or "侧" in note_upper or "边" in model_upper or "侧" in model_upper:
                return True
        else:
            if required in note_upper or required in model_upper:
                return True

        # 兜底逻辑：只有当以下条件全部满足时才视为通用刀模
        # 1. 刀型号号中没有"前"、"边"、"后"（排除本身就有部位标识的刀模）
        # 2. 数量 >= 3（说明是通用刀模，不是特定用途）
        # 3. 备注为空或备注中没有"前"、"边"、"后"
        model_has_dir = any(k in model_upper for k in ['前', '边', '后', '侧'])
        note_has_dir = any(k in note_upper for k in ['前', '边', '后', '侧'])
        total_qty = rec.get('total_quantity', 0) or rec.get('quantity', 0)

        if not model_has_dir and total_qty >= 3 and not note_has_dir:
            return True  # 视为通用刀模，可匹配任何类型

        return False
    # 获取工艺单的段数
    process_segment_count = count_model_segments(normalized_model)

    # 收集所有匹配结果
    matches = []

    # 1. 提取前缀
    prefix = _extract_model_prefix(normalized_model)

    # 2. 优先在前缀分组内查找
    if prefix and prefix in nonv_prefix_index:
        prefix_group = nonv_prefix_index[prefix]

        # 2.1 直接匹配标准化型号（同时要求段数相同和关键字匹配）
        for rec in prefix_group:
            if rec['model'] == normalized_model and note_has_keyword(rec, keyword_required):
                matches.append(rec)

        # 2.2 使用 get_die_mold_keys 在该分组内匹配（要求段数相同和关键字匹配）
        candidate_keys = get_die_mold_keys(normalized_model)
        for rec in prefix_group:
            if rec in matches:
                continue
            # 段数必须相同
            if count_model_segments(rec['model']) != process_segment_count:
                continue
            rec_keys = get_die_mold_keys(rec['model'])
            if candidate_keys & rec_keys:  # 集合交集
                # 关键字匹配检查
                if note_has_keyword(rec, keyword_required):
                    matches.append(rec)

    # 3. 后备方案：使用完整键索引（处理前缀不匹配的情况，要求段数相同和关键字匹配）
    candidate_keys = get_die_mold_keys(normalized_model)
    for key in candidate_keys:
        if key in nonv_full_index:
            rec = nonv_full_index[key]
            if rec in matches:
                continue
            # 段数必须相同
            if count_model_segments(rec['model']) == process_segment_count:
                # 关键字匹配检查
                if note_has_keyword(rec, keyword_required):
                    matches.append(rec)

    # 4. 第二次匹配：如果仍未匹配到，尝试去掉特殊后缀/括号内容后重试
    #    用于匹配刀模型号无特殊标注但备注栏有对应标注的库存记录
    #    例如：A4K5-AB-58-011双针前 ↔ A4K5-AB-58-011 + 备注"双针前"
    #         E1K2-C-57-009(EVA) ↔ E1K2-C-57-009 + 备注"EVA"
    #    注意：此时备注栏必须包含被去掉的特殊内容
    if not matches:
        # 提取特殊后缀/括号内容
        special_suffixes = []
        # 提取括号内容如 (EVA)、（EVA）
        bracket_match = re.search(r'[（(]([^）)]+)[)）]$', normalized_model)
        if bracket_match:
            special_suffixes.append(bracket_match.group(1))
        # 提取双针前等特殊后缀
        for suffix in ("双针前", "双针后", "双针边"):
            if suffix in normalized_model:
                special_suffixes.append(suffix)

        for special in special_suffixes:
            normalized_second = normalized_model.replace(special, "")
            if normalized_second:
                prefix_second = _extract_model_prefix(normalized_second)
                if prefix_second and prefix_second in nonv_prefix_index:
                    prefix_group = nonv_prefix_index[prefix_second]
                    candidate_keys_second = get_die_mold_keys(normalized_second)
                    for rec in prefix_group:
                        if rec in matches:
                            continue
                        if count_model_segments(rec['model']) != count_model_segments(normalized_second):
                            continue
                        rec_keys = get_die_mold_keys(rec['model'])
                        if candidate_keys_second & rec_keys:
                            # 第二次匹配时备注栏必须包含特殊内容
                            note = rec.get('note', '') or ''
                            if special in note:
                                matches.append(rec)
                # 也检查完整键索引
                if not matches:
                    for key in get_die_mold_keys(normalized_second):
                        if key in nonv_full_index:
                            rec = nonv_full_index[key]
                            if rec in matches:
                                continue
                            if count_model_segments(rec['model']) == count_model_segments(normalized_second):
                                note = rec.get('note', '') or ''
                                if special in note:
                                    matches.append(rec)

    return matches


def _match_v_mold_fast(model_str, v_models):
    """快速匹配V型刀模（优化版）

    返回所有匹配的记录列表。

    匹配规则：段数必须相同
    - V19舌上（1段）只能匹配1段的库存刀模
    - V19舌上-A-023（3段）只能匹配3段的库存刀模
    """
    model_str = model_str.strip()
    model_segment_count = count_model_segments(model_str)
    matches = []

    # 1. 直接匹配（段数相同）
    for rec in v_models:
        if model_str == rec['model'] and count_model_segments(rec['model']) == model_segment_count:
            matches.append(rec)

    # 2. 包含匹配（双向，段数必须相同）
    # 注意：不能简单包含匹配，要确保关键字完整
    # 例如：V19下 不能匹配 V19皮革上下（因为下前面有关键字差异）
    for rec in v_models:
        if rec in matches:
            continue
        inv_model = rec['model']
        if count_model_segments(inv_model) != model_segment_count:
            continue
        # 直接包含检查（适用于 V19A 包含 V19 这种简单情况）
        if model_str in inv_model or inv_model in model_str:
            matches.append(rec)

    # 3. 处理上下组合型（段数相同）
    if "上下" in model_str:
        base_model = model_str.replace("上下", "").replace("V", "")
        for rec in v_models:
            if rec in matches:
                continue
            if count_model_segments(rec['model']) != model_segment_count:
                continue
            inv_base = rec['model'].replace("上下", "").replace("V", "")
            if base_model == inv_base:
                matches.append(rec)

    # 4. 处理舌字变体（段数相同）
    for rec in v_models:
        if rec in matches:
            continue
        inv_model = rec['model']
        if count_model_segments(inv_model) != model_segment_count:
            continue
        # 去掉舌字后匹配
        model_no_tongue = model_str.replace('舌', '')
        inv_no_tongue = inv_model.replace('舌', '')
        if model_no_tongue == inv_no_tongue:
            matches.append(rec)
            continue
        # 上/下匹配
        if model_no_tongue.endswith('上') and inv_no_tongue == model_no_tongue[:-1] + '舌上':
            matches.append(rec)
        if model_no_tongue.endswith('下') and inv_no_tongue == model_no_tongue[:-1] + '舌下':
            matches.append(rec)

    # 5. 处理舌上/舌下 匹配（段数相同）
    # 关键字规则：有舌时舌前是整体+舌+上下，无舌时上下前+上下
    # V19舌下(舌前=V19) ≠ V19皮革上下(舌前=V19皮革)
    for rec in v_models:
        if rec in matches:
            continue
        inv_model = rec['model']
        if count_model_segments(inv_model) != model_segment_count:
            continue
        inv_normalized = inv_model.replace('、', '').replace(',', '').replace('，', '')
        model_normalized = model_str.replace('、', '').replace(',', '').replace('，', '')

        import re
        v_match = re.match(r'V(\d+)', model_normalized)
        inv_v_match = re.match(r'V(\d+)', inv_normalized)
        if v_match and inv_v_match and v_match.group(1) == inv_v_match.group(1):
            model_part = model_normalized.replace(v_match.group(0), '')
            inv_part = inv_normalized.replace(inv_v_match.group(0), '')

            # 提取舌前部分（整体作为关键字）
            if '舌' in model_part:
                idx = model_part.index('舌')
                model_before_tongue = model_part[:idx]  # 舌前整体
            else:
                m = re.search(r'[上下]', model_part)
                model_before_tongue = model_part[:m.start()] if m else ''

            if '舌' in inv_part:
                idx = inv_part.index('舌')
                inv_before_tongue = inv_part[:idx]
            else:
                m = re.search(r'[上下]', inv_part)
                inv_before_tongue = inv_part[:m.start()] if m else ''

            # 舌前整体必须相同
            if model_before_tongue != inv_before_tongue:
                continue

            # 检查部位
            model_has_upper = '上' in model_part
            model_has_lower = '下' in model_part
            inv_has_upper = '上' in inv_part
            inv_has_lower = '下' in inv_part

            if model_has_upper and not model_has_lower:
                if inv_has_upper:
                    matches.append(rec)
            if model_has_lower and not model_has_upper:
                if inv_has_lower:
                    matches.append(rec)
            if model_has_upper and model_has_lower:
                if inv_has_upper and inv_has_lower:
                    matches.append(rec)

    return matches


def check_mold_in_stock(model, type_name, mold_type, inventory_models):
    """检查刀模是否在库存中"""
    # 标准化处理
    normalized = normalize_process_ref(model) if mold_type == "非V型" else model.strip()
    keys = get_die_mold_keys(normalized) if mold_type == "非V型" else {model.strip()}

    found_rec = None
    for inv_model, rec in inventory_models.items():
        if mold_type == "V型":
            # V型刀模匹配
            if model.strip() in inv_model or inv_model in model.strip():
                found_rec = rec
                break
            # 处理上下组合型
            if "上下" in model or "上下" in inv_model:
                base1 = model.replace("上下", "").replace("V", "")
                base2 = inv_model.replace("上下", "").replace("V", "")
                if base1 == base2:
                    found_rec = rec
                    break
        else:
            # 非V型刀模匹配
            inv_keys = get_die_mold_keys(inv_model)
            if keys & inv_keys:
                found_rec = rec
                break

    if found_rec:
        return {
            "model": model,
            "inventory_model": found_rec.get("model"),
            "type_name": type_name,
            "mold_type": mold_type,
            "in_stock": True,
            "position": found_rec.get("position"),
            "quantity": found_rec.get("available", found_rec.get("quantity")),
            "total_quantity": found_rec.get("total_quantity", found_rec.get("quantity")),
            "borrowed": found_rec.get("borrowed", 0),
            "available": found_rec.get("available", found_rec.get("quantity")),
            "note": found_rec.get("note"),
            "borrower": found_rec.get("borrower"),
            "is_borrowed": found_rec.get("is_borrowed", False),
            "borrow_date": found_rec.get("borrow_date", ""),
        }
    else:
        return {
            "model": model,
            "inventory_model": None,
            "type_name": type_name,
            "mold_type": mold_type,
            "in_stock": False,
            "position": None,
            "quantity": None,
            "total_quantity": None,
            "borrowed": None,
            "available": None,
            "note": None,
            "borrower": None,
            "is_borrowed": False,
            "borrow_date": "",
        }


# ── API: 批量添加刀模 ─────────────────────────────────
@app.route('/api/die_molds/purchase', methods=['POST'])
def api_purchase_die_mold():
    """购买刀模：将缺失的刀模添加到刀模汇总表
    
    请求体：
    {
        "molds": [
            {
                "model": "V23舌上",
                "type_name": "上眉",
                "mold_type": "V型"
            },
            {
                "model": "A1-125#-A-59CM前片",
                "type_name": "前片",
                "mold_type": "非V型"
            }
        ]
    }
    """
    data = request.get_json()
    molds = data.get('molds', [])
    
    if not molds:
        return jsonify({"error": "没有要购买的刀模"}), 400
    
    # 读取现有库存
    inventory = ml_load_die_molds(DIE_MOLD_FILE, DIE_MOLD_SHEET)
    existing_models = {rec['model'] for rec in inventory}
    
    added = 0
    
    for mold in molds:
        model = mold.get('model', '').strip()
        mold_type = mold.get('mold_type', '')
        type_name = mold.get('type_name', '')
        
        if not model:
            continue
        
        # 检查是否已存在（简化匹配）
        if any(model in existing or existing in model for existing in existing_models):
            continue
        
        if mold_type == 'V型':
            # V型刀模（上眉、下眉）→ 刀模型号列
            record = {
                "position": mold.get('position', ''),
                "model": model,
                "mold_type": "V型",
                "total_quantity": 1,
                "borrowed": 0,
                "available": 1,
                "note": type_name,  # 备注放类型名（上眉/下眉）
            }
        else:
            # 非V型刀模（前片、边片、后片）
            # 从工艺单型号中提取纯净型号（去掉片名后缀）
            clean_model = normalize_process_ref(model) or model
            record = {
                "position": mold.get('position', ''),
                "model": clean_model,  # 刀模型号列放纯净型号
                "mold_type": "非V型",
                "total_quantity": 1,
                "borrowed": 0,
                "available": 1,
                "note": type_name,  # 备注列放类型名（前片/边片/后片）
            }
        
        # 生成ID（使用行号作为ID）
        record_id = len(inventory) + 1
        record['id'] = record_id
        record['row'] = len(inventory) + 2
        inventory.append(record)
        existing_models.add(record['model'])
        added += 1
    
    # 保存
    from utils.excel_reader import save_die_molds
    save_die_molds(inventory)
    
    # 清除缓存
    invalidate_inventory_cache()
    
    return jsonify({
        "success": True,
        "added": added,
        "message": f"成功购买 {added} 个刀模"
    })


@app.route('/api/die_molds/batch', methods=['POST'])
def api_batch_add_die_molds():
    """批量添加刀模到库存"""
    data = request.get_json()
    molds = data.get('molds', [])
    action = data.get('action', '')

    if not molds:
        return jsonify({"error": "没有要添加的刀模"}), 400

    # 读取现有库存
    inventory = ml_load_die_molds(DIE_MOLD_FILE, DIE_MOLD_SHEET)
    existing_models = {rec['model'] for rec in inventory}

    added = 0

    for mold in molds:
        model = mold.get('model', '').strip()
        mold_type = mold.get('mold_type', '')
        type_name = mold.get('type_name', '')

        if not model:
            continue

        # 检查是否已存在（简化匹配）
        if any(model in existing or existing in model for existing in existing_models):
            continue

        if mold_type == 'V型':
            # V型刀模（上眉、下眉）→ 刀模型号列
            record = {
                "position": "",
                "model": model,
                "mold_type": "V型",
                "total_quantity": 1,
                "borrowed": 0,
                "available": 1,
                "note": type_name,  # 备注放类型名（上眉/下眉）
            }
        else:
            # 非V型刀模（前片、边片、后片）→ 刀模型号列放型号，备注放类型名
            record = {
                "position": "",
                "model": model,
                "mold_type": "非V型",
                "total_quantity": 1,
                "borrowed": 0,
                "available": 1,
                "note": type_name,  # 备注放类型名（前片/边片/后片）
            }

        # 生成ID（使用行号作为ID）
        record_id = len(inventory) + 1
        record['id'] = record_id
        record['row'] = len(inventory) + 2
        inventory.append(record)
        existing_models.add(model)
        added += 1

    # 保存
    from utils.excel_reader import save_die_molds
    save_die_molds(inventory)
    
    # 清除缓存
    invalidate_inventory_cache()

    return jsonify({
        "success": True,
        "added": added,
        "message": f"成功添加 {added} 个刀模到库存"
    })


# ── API: 刀模 CRUD ─────────────────────────────────────
@app.route('/api/die_molds', methods=['GET'])
def api_get_die_molds():
    """获取刀模列表，支持搜索"""
    keyword = request.args.get('q', '')
    records = search_die_molds(keyword) if keyword else load_die_molds()
    return jsonify(records)


@app.route('/api/die_molds', methods=['POST'])
def api_add_die_mold():
    """新增刀模"""
    data = request.get_json()
    total_qty = int(data.get('total_quantity', data.get('quantity', 1)))
    borrowed = int(data.get('borrowed', 0))
    record = {
        "position": data.get('position', ''),
        "model": data.get('model', ''),
        "mold_type": data.get('mold_type', ''),
        "total_quantity": total_qty,
        "borrowed": borrowed,
        "available": total_qty - borrowed,
        "note": data.get('note', ''),
    }
    result = add_die_mold(record)
    # 清除缓存
    invalidate_inventory_cache()
    return jsonify(result)


@app.route('/api/die_molds/<int:id>', methods=['PUT'])
def api_update_die_mold(id):
    """更新刀模"""
    data = request.get_json()
    total_qty = int(data.get('total_quantity', data.get('quantity', 1)))
    borrowed = int(data.get('borrowed', 0))
    record = {
        "position": data.get('position', ''),
        "model": data.get('model', ''),
        "mold_type": data.get('mold_type', ''),
        "total_quantity": total_qty,
        "borrowed": borrowed,
        "available": total_qty - borrowed,
        "note": data.get('note', ''),
    }
    result = update_die_mold(id, record)
    if result:
        # 清除缓存
        invalidate_inventory_cache()
        return jsonify(result)
    return jsonify({"error": "记录不存在"}), 404


@app.route('/api/die_molds/<int:id>', methods=['DELETE'])
def api_delete_die_mold(id):
    """删除刀模"""
    delete_die_mold(id)
    # 清除缓存
    invalidate_inventory_cache()
    return jsonify({"success": True})


@app.route('/api/die_molds/<int:id>/offline', methods=['POST'])
def api_offline_die_mold(id):
    """下架刀模：将记录移动到下架工作表"""
    result = offline_die_mold(id)
    if result:
        # 清除缓存
        invalidate_inventory_cache()
        return jsonify({"success": True, "record": result})
    return jsonify({"error": "记录不存在"}), 404


@app.route('/api/die_molds/offline', methods=['GET'])
def api_get_offline_die_molds():
    """获取下架工作表中的刀模"""
    if not DIE_MOLD_FILE.exists():
        return jsonify([])

    wb = openpyxl.load_workbook(DIE_MOLD_FILE, data_only=True)
    if "下架" not in wb.sheetnames:
        wb.close()
        return jsonify([])

    ws = wb["下架"]
    records = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or not row[1]:
            continue
        records.append({
            "id": i,
            "position": row[0] or "",
            "model": str(row[1]).strip() if row[1] else "",
            "borrower": row[2] or "",
            "quantity": row[3] or 0,
            "note": row[4] or "",
        })
    wb.close()
    return jsonify(records)


# ── API: 完整库存（含借出状态）──────────────────────────
@app.route('/api/inventory/full', methods=['GET'])
def api_get_full_inventory():
    """获取完整库存列表（含借出状态标注）

    返回刀模汇总表中的所有记录，并标注：
    - 是否有借出（从 borrow_records.db 查询）
    - 借出人、借出日期等信息
    """
    # 获取所有刀模记录
    records = load_die_molds()

    # 获取当前借出状态汇总（按型号+位置）
    borrow_summary = get_borrow_summary_by_model_position()

    # 关联借出信息
    result = []
    for rec in records:
        model = rec.get('model', '')
        position = rec.get('position', '')

        # 查找对应的借出记录
        borrow_info = borrow_summary.get((model, position), {})

        result.append({
            "id": rec.get('id'),
            "position": position,
            "model": model,
            "mold_type": rec.get('mold_type', ''),
            "total_quantity": rec.get('total_quantity', 0),
            "borrowed": rec.get('borrowed', 0),
            "available": rec.get('available', 0),
            "note": rec.get('note', ''),
            # 借出状态
            "is_borrowed": bool(borrow_info),
            "borrower": borrow_info.get('borrower', ''),
            "borrow_date": borrow_info.get('borrow_date', ''),
        })

    return jsonify(result)


# ── API: 刀模搜索（用于领用）────────────────────────────
@app.route('/api/die_molds/search_for_borrow', methods=['GET'])
def api_search_die_molds_for_borrow():
    """搜索刀模用于领用填写

    根据输入的关键词搜索刀模，返回所有匹配的刀模列表（包括位置信息）
    **只返回可用库存 > 0 的刀模**

    查询参数：
    - q: 搜索关键词（刀模型号的一部分）
    - limit: 返回数量限制，默认50
    """
    keyword = request.args.get('q', '').strip()
    limit = int(request.args.get('limit', 50))

    if not keyword:
        return jsonify([])

    # 获取所有刀模记录
    records = load_die_molds()

    # 搜索匹配的记录（关键词匹配刀模型号或位置）
    # 只返回 available > 0 的刀模
    keyword_lower = keyword.lower()
    matches = []
    for rec in records:
        available = rec.get('available', 0)
        # 跳过可用库存为0的刀模
        if available <= 0:
            continue

        model = rec.get('model', '').lower()
        position = str(rec.get('position', '')).lower()
        if keyword_lower in model or keyword_lower in position:
            matches.append({
                "id": rec.get('id'),
                "model": rec.get('model'),
                "position": rec.get('position'),
                "mold_type": rec.get('mold_type'),
                "total_quantity": rec.get('total_quantity'),
                "borrowed": rec.get('borrowed'),
                "available": available,
                "note": rec.get('note'),
            })
            if len(matches) >= limit:
                break

    return jsonify(matches)


# ── API: 刀模搜索（通用实时搜索）────────────────────────
@app.route('/api/die_molds/search', methods=['GET'])
def api_search_die_molds():
    """刀模实时搜索 - 模糊匹配

    模糊匹配规则：搜索词必须包含在刀模型号中（不区分大小写）
    例如：搜索 "A1-502" 会匹配 "A1-502-58"、"A1-502-58-066" 等

    查询参数：
    - q: 搜索关键词（刀模型号的一部分）
    - limit: 返回数量限制，默认30
    """
    keyword = request.args.get('q', '').strip()
    limit = int(request.args.get('limit', 30))

    if not keyword:
        return jsonify([])

    # 获取库存缓存
    cache = get_inventory_cache()
    records = cache.get('records', [])

    if not records:
        records = load_die_molds()

    # 模糊匹配：搜索词包含在刀模型号中（去掉括号后比较，因为括号是说明不是另一个刀模）
    keyword_lower = keyword.lower()
    keyword_stripped = _REVISION_RE.sub("", keyword).strip().lower()
    matches = []
    for rec in records:
        model = rec.get('model', '')
        position = str(rec.get('position', ''))
        model_stripped = _REVISION_RE.sub("", model).strip()

        # 模糊匹配：搜索词或去括号后的搜索词包含在去括号后的刀模型号中
        if keyword_lower in model.lower() or (keyword_stripped and keyword_stripped in model_stripped.lower()):
            matches.append({
                "model": model,
                "position": position,
                "mold_type": rec.get('mold_type', ''),
                "total_quantity": rec.get('total_quantity', 0),
                "borrowed": rec.get('borrowed', 0),
                "available": rec.get('available', 0),
                "note": rec.get('note', ''),
            })
            if len(matches) >= limit:
                break

    return jsonify(matches)


# ── API: 刀模搜索（精准匹配）────────────────────────────
@app.route('/api/die_molds/search_exact', methods=['GET'])
def api_search_die_molds_exact():
    """刀模精准搜索 - 段数相同且搜索词与模型号完全一致

    用于最终查询按钮，结果必须与输入的搜索词完全匹配（不区分大小写）

    查询参数：
    - q: 搜索关键词（刀模型号）
    - limit: 返回数量限制，默认100
    """
    keyword = request.args.get('q', '').strip()
    limit = int(request.args.get('limit', 100))

    if not keyword:
        return jsonify([])

    # 获取库存缓存
    cache = get_inventory_cache()
    records = cache.get('records', [])

    if not records:
        records = load_die_molds()

    # 精准匹配：
    # - V型：候选集合匹配（简写扩展：下↔舌下↔舌上、下↔上下；材质是关键字不能混）
    # - 棚帽（含舌芯/包边）：只看病号是否含这些词，不管方位
    # - 非V型：去掉括号精确匹配（部位标识如前/后/边/侧是刀号的一部分，不剥离）
    keyword_lower = keyword.lower()
    keyword_stripped = _REVISION_RE.sub("", keyword).strip()
    is_v_keyword = keyword_stripped.startswith('V') if keyword_stripped else False
    is_pengmao = '舌芯' in keyword_stripped or '包边' in keyword_stripped

    # 构建 V 型搜索词的候选集合
    v_candidates = set()
    if is_v_keyword and not is_pengmao:
        v_candidates.add(keyword_stripped.lower())
        v_match = re.match(r'^(V\d+)', keyword_stripped)
        if v_match:
            v_prefix = v_match.group(1)
            rest = keyword_stripped[len(v_prefix):]
            has_upper = '上' in rest
            has_lower = '下' in rest
            if has_upper and has_lower:
                base = rest.replace('上下', '').replace('、', '')
                for suffix in ['上', '下', '舌上', '舌下', '舌上、下', '上下']:
                    v_candidates.add(f"{v_prefix}{base}{suffix}".lower())
            elif has_upper:
                base = rest.replace('上', '')
                for suffix in ['上', '舌上', '舌上、下', '上下']:
                    v_candidates.add(f"{v_prefix}{base}{suffix}".lower())
            elif has_lower:
                base = rest.replace('下', '')
                for suffix in ['下', '舌下', '舌上、下', '上下']:
                    v_candidates.add(f"{v_prefix}{base}{suffix}".lower())
            else:
                # 只有V编号，无方位
                pass

    # 棚帽模式也需要生成候选集合
    if is_v_keyword and is_pengmao:
        v_match = re.match(r'^(V\d+)', keyword_stripped)
        if v_match:
            v_prefix = v_match.group(1)
            rest = keyword_stripped[len(v_prefix):]
            # 提取包边/舌芯部分
            has_baobian = '包边' in rest
            has_shexin = '舌芯' in rest
            mold_type = '包边' if has_baobian else '舌芯'
            # 去掉包边/舌芯后看方位
            rest_without_type = rest.replace('包边', '').replace('舌芯', '')
            has_upper = '上' in rest_without_type
            has_lower = '下' in rest_without_type

            if has_upper and has_lower:
                base = rest_without_type.replace('上下', '').replace('、', '')
                for suffix in ['上', '下', '舌上', '舌下', '舌上、下', '上下']:
                    v_candidates.add(f"{v_prefix}{mold_type}{base}{suffix}".lower())
            elif has_upper:
                base = rest_without_type.replace('上', '')
                for suffix in ['上', '舌上', '舌上、下', '上下']:
                    v_candidates.add(f"{v_prefix}{mold_type}{base}{suffix}".lower())
            elif has_lower:
                base = rest_without_type.replace('下', '')
                for suffix in ['下', '舌下', '舌上、下', '上下']:
                    v_candidates.add(f"{v_prefix}{mold_type}{base}{suffix}".lower())

    matches = []
    for rec in records:
        model = rec.get('model', '')
        model_stripped = _REVISION_RE.sub("", model).strip()
        model_stripped_lower = model_stripped.lower()
        note = rec.get('note', '') or ''

        matched = False
        if is_pengmao:
            # 棚帽（V型的一种）：候选集合匹配 + V编号前缀检查
            if model_stripped_lower in v_candidates:
                v_match = re.match(r'^(V\d+)', keyword_stripped)
                if v_match:
                    v_prefix = v_match.group(1).lower()
                    if model_stripped_lower.startswith(v_prefix):
                        matched = True
        elif is_v_keyword:
            # V型：候选集合匹配
            if model_stripped_lower in v_candidates:
                matched = True
        else:
            # 非V型：去掉括号精确匹配 + 备注部位关键字匹配
            # 精确匹配优先，其次用搜索词中的部位关键字（前/后/边/侧）匹配库存备注
            # 搜索词含位置关键字时，分离出基础型号和位置部分
            # 例如：搜索"A4K5-AB-58-011双针前" → 基础型号="A4K5-AB-58-011"，位置="双针前"
            base_keyword = keyword_stripped
            location_in_kw = None
            # 先检查复合位置词（双针前/双针后等），再检查单个位置字
            for d in ['双针前', '双针后', '前片', '后片', '边片', '前', '后', '边', '侧']:
                if d in keyword_stripped:
                    location_in_kw = d
                    base_keyword = keyword_stripped.replace(d, '').strip()
                    break

            # 精确匹配：基础型号与model一致（或搜索词整体与model一致）
            exact_match = model.lower() == keyword_lower or (keyword_stripped and model_stripped_lower == keyword_stripped.lower())
            # 基础型号匹配：去掉位置关键字后的搜索词与model一致
            base_match = base_keyword.lower() == model_stripped_lower

            if exact_match:
                # 精确匹配：model与搜索词完全一致（包括搜索词含位置关键字的情况）
                # 例如：搜索"A4K5-AB-58-011双针前"，库存model为"A4K5-AB-58-011双针前"
                # 或者：搜索"A4K5-AB-58-011双针前"，库存model为"A4K5-AB-58-011"（备注含"双针前"）
                matched = True
            elif base_match and location_in_kw:
                # 基础型号匹配且搜索词含位置关键字，检查备注是否含该位置
                # 例如：搜索"A4K5-AB-58-011双针前"，库存model为"A4K5-AB-58-011"（备注含"双针前"）
                note_u = note.upper()
                if location_in_kw == '双针前' and '双针前' in note_u:
                    matched = True
                elif location_in_kw == '双针后' and '双针后' in note_u:
                    matched = True
                else:
                    # 单字方向：前/后/边/侧
                    dir_mapping = {'前片': '前', '后片': '后', '边片': '边'}
                    actual_dir = dir_mapping.get(location_in_kw, location_in_kw)
                    if actual_dir == '前' and '前' in note_u:
                        matched = True
                    elif actual_dir == '后' and '后' in note_u:
                        matched = True
                    elif actual_dir in ('边', '侧') and ('边' in note_u or '侧' in note_u):
                        matched = True

        if matched:
            matches.append({
                "model": model,
                "position": str(rec.get('position', '')),
                "mold_type": rec.get('mold_type', ''),
                "total_quantity": rec.get('total_quantity', 0),
                "borrowed": rec.get('borrowed', 0),
                "available": rec.get('available', 0),
                "note": note,
            })
            if len(matches) >= limit:
                break

    return jsonify(matches)


# ── API: 位置搜索（用于库存管理）────────────────────────
@app.route('/api/positions/search', methods=['GET'])
def api_search_positions():
    """按位置搜索刀模，返回位置列表及每个位置的刀模数量

    查询参数：
    - q: 搜索关键词（位置的一部分）
    - limit: 返回数量限制，默认50
    """
    keyword = request.args.get('q', '').strip()
    limit = int(request.args.get('limit', 50))

    if not keyword:
        return jsonify([])

    # 获取库存缓存
    cache = get_inventory_cache()
    records = cache.get('records', [])

    if not records:
        records = load_die_molds()

    # 后端位置过滤
    keyword_lower = keyword.lower()
    position_count = {}

    for rec in records:
        position = str(rec.get('position', ''))
        if keyword_lower in position.lower():
            if position not in position_count:
                position_count[position] = {"count": 0, "records": []}
            position_count[position]["count"] += 1
            position_count[position]["records"].append({
                "model": rec.get('model', ''),
                "position": position,
                "available": rec.get('available', 0),
            })

    # 转换为列表并限制数量
    result = []
    for pos, data in position_count.items():
        result.append({
            "position": pos,
            "count": data["count"],
            "sample": data["records"][:3]  # 只返回前3个样本
        })
        if len(result) >= limit:
            break

    return jsonify(result)


@app.route('/api/match', methods=['POST'])
def api_match():
    """上传工艺单并匹配"""
    if 'file' not in request.files:
        return jsonify({"error": "没有上传文件"}), 400

    file = request.files['file']
    if not file.filename.endswith('.xlsx'):
        return jsonify({"error": "只支持 .xlsx 文件"}), 400

    # 保存上传文件
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    temp_file = BASE_DIR / f"temp_process_{timestamp}.xlsx"
    file.save(temp_file)

    try:
        # 读取工艺单
        count_nonv, count_v, fc, rc = ml_load_process_refs(BASE_DIR, f"temp_process_{timestamp}.xlsx", "Sheet1")

        # 读取库存
        records = ml_load_die_molds(DIE_MOLD_FILE, DIE_MOLD_SHEET)

        # 执行匹配
        matched = ml_match_die_molds(records, count_nonv, count_v)

        # 分类结果
        existing = []
        missing = []
        borrowed = []

        # 收集工艺单中的所有引用
        all_refs = set(count_nonv.keys()) | set(count_v.keys())

        for rec in matched:
            if rec.get('borrower') and rec['borrower'].strip():
                borrowed.append({
                    "model": rec['model'],
                    "position": rec['position'],
                    "borrower": rec['borrower'],
                    "type": get_mold_type(rec['model'])
                })
            else:
                existing.append({
                    "model": rec['model'],
                    "position": rec['position'],
                    "quantity": rec['quantity'],
                    "type": get_mold_type(rec['model'])
                })

        # 找出缺失的刀模（工艺单中需要但库存没有的）
        inventory_models = {rec['model'] for rec in records}
        for ref in all_refs:
            # 检查是否是V型
            is_v = False
            for v_key in count_v.keys():
                if ref == v_key or ref in v_key or v_key in ref:
                    is_v = True
                    break

            # 检查是否在库存中
            found = False
            for inv_model in inventory_models:
                if is_v:
                    if ref in inv_model or inv_model in ref:
                        found = True
                        break
                else:
                    keys = get_die_mold_keys(ref)
                    if any(k in inventory_models for k in keys):
                        found = True
                        break

            if not found:
                mold_type = "V型" if is_v else "非V型"
                missing.append({
                    "model": ref,
                    "type": mold_type,
                    "quantity": 1
                })

        # 去重
        seen = set()
        unique_missing = []
        for m in missing:
            if m['model'] not in seen:
                seen.add(m['model'])
                unique_missing.append(m)

        return jsonify({
            "file_count": fc,
            "row_count": rc,
            "existing": existing,
            "missing": unique_missing,
            "borrowed": borrowed,
            "summary": {
                "total_existing": len(existing),
                "total_missing": len(unique_missing),
                "total_borrowed": len(borrowed)
            }
        })

    finally:
        # 清理临时文件
        if temp_file.exists():
            temp_file.unlink()


def get_mold_type(model):
    """获取刀模类型"""
    if is_v_type(model):
        return "V型"
    return "非V型"


# ── API: 刀模领用归还 ──────────────────────────────────
def _update_die_mold_borrow_status(mold_model, mold_position, quantity, is_return=False, borrower="", borrow_date=""):
    """更新刀模汇总表中的借出状态

    mold_model: 刀模型号
    mold_position: 刀模位置
    quantity: 借出/归还数量
    is_return: True 表示归还
    borrower: 借出人（借出时必填）
    borrow_date: 借出日期（借出时必填）
    """
    import re
    mold_model = mold_model.strip()
    mold_position = str(mold_position).strip() if mold_position else ''
    if not mold_model:
        return

    records = load_die_molds()  # 使用规范化结构的 load_die_molds
    updated = False

    for rec in records:
        if rec.get('model', '').strip() == mold_model and str(rec.get('position', '')).strip() == mold_position:
            # 更新已借出和可用库存
            if is_return:
                # 归还：增加可用，减少已借出
                rec['borrowed'] = max(0, rec.get('borrowed', 0) - quantity)
                rec['available'] = rec.get('total_quantity', 0) - rec['borrowed']
                # 归还时只清除备注中对应的借出标注
                note = rec.get('note', '') or ''
                # 精确匹配：只删除特定借出人的标注
                pattern = rf'\[借出:{re.escape(borrower)} {re.escape(borrow_date)}\]'
                note = re.sub(pattern, '', note).strip()
                # 清理多余的空格
                note = re.sub(r'\s+', ' ', note).strip()
                rec['note'] = note
            else:
                # 借出：减少可用，增加已借出
                rec['borrowed'] = rec.get('borrowed', 0) + quantity
                rec['available'] = rec.get('total_quantity', 0) - rec['borrowed']
                # 借出时在备注中追加借出标注（不覆盖已有的）
                note = rec.get('note', '') or ''
                borrow_tag = f"[借出:{borrower} {borrow_date}]"
                # 追加新的借出标注
                rec['note'] = f"{note} {borrow_tag}".strip()
            updated = True
            break

    if updated:
        from utils.excel_reader import save_die_molds
        save_die_molds(records)



@app.route('/api/borrow/init', methods=['POST'])
def api_init_borrow_db():
    """初始化领用数据库"""
    init_borrow_db()
    return jsonify({"success": True, "message": "领用数据库初始化完成"})


@app.route('/api/borrow/records', methods=['GET'])
def api_get_borrow_records():
    """获取领用记录列表"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status = request.args.get('status')
    borrower = request.args.get('borrower')

    records = get_borrow_records(start_date, end_date, status, borrower)
    return jsonify({"success": True, "records": records})


@app.route('/api/borrow/record/<int:record_id>', methods=['GET'])
def api_get_borrow_record(record_id):
    """获取单条领用记录"""
    record = get_borrow_record_by_id(record_id)
    if record:
        return jsonify({"success": True, "record": record})
    return jsonify({"success": False, "error": "记录不存在"}), 404


@app.route('/api/borrow/record', methods=['POST'])
def api_add_borrow_record():
    """添加领用记录"""
    data = request.get_json()

    mold_model = data.get('mold_model', '').strip()
    mold_type = data.get('mold_type', '').strip()  # 部位：前片/边片/后片/V型
    borrower = data.get('borrower', '').strip()
    borrow_date = data.get('borrow_date', '').strip()
    quantity = int(data.get('quantity', 1))
    note = data.get('note', '').strip()
    mold_position = data.get('mold_position', '').strip()

    if not mold_model or not borrower or not borrow_date:
        return jsonify({"success": False, "error": "刀模型号、领用人、领用日期不能为空"}), 400

    # 检查库存是否足够
    records = load_die_molds()
    mold_found = None
    for rec in records:
        if rec.get('model', '').strip() == mold_model and str(rec.get('position', '')).strip() == mold_position:
            mold_found = rec
            break

    if not mold_found:
        return jsonify({"success": False, "error": f"未找到刀模 {mold_model}，位置 {mold_position}"}), 400

    available = mold_found.get('available', 0)
    if available < quantity:
        return jsonify({"success": False, "error": f"库存不足！当前可用数量: {available}，申请数量: {quantity}"}), 400

    record_id = add_borrow_record(mold_model, mold_type, borrower, borrow_date, quantity, note, mold_position)

    # 更新刀模汇总表中的借出状态
    _update_die_mold_borrow_status(mold_model, mold_position, quantity, is_return=False, borrower=borrower, borrow_date=borrow_date)

    invalidate_inventory_cache()

    return jsonify({"success": True, "record_id": record_id, "message": "领用记录添加成功"})


@app.route('/api/borrow/record/<int:record_id>/return', methods=['POST'])
def api_return_mold(record_id):
    """归还刀模"""
    data = request.get_json()
    return_date = data.get('return_date', '').strip()
    returner = data.get('returner', '').strip()
    note = data.get('note', '').strip()

    if not return_date:
        return jsonify({"success": False, "error": "归还日期不能为空"}), 400

    # 先获取记录信息，以便更新刀模汇总表
    borrow_record = get_borrow_record_by_id(record_id)
    if borrow_record:
        mold_model = borrow_record.get('mold_model', '')
        mold_position = borrow_record.get('mold_position', '')
        quantity = borrow_record.get('quantity', 1)
        borrower = borrow_record.get('borrower', '')
        borrow_date = borrow_record.get('borrow_date', '')
        # 更新刀模汇总表中的借出状态
        _update_die_mold_borrow_status(mold_model, mold_position, quantity, is_return=True, borrower=borrower, borrow_date=borrow_date)

    return_mold(record_id, return_date, returner, note)
    invalidate_inventory_cache()

    return jsonify({"success": True, "message": "归还成功"})


@app.route('/api/borrow/record/<int:record_id>', methods=['DELETE'])
def api_delete_borrow_record(record_id):
    """删除领用记录（同时归还刀模）"""
    # 获取记录信息
    borrow_record = get_borrow_record_by_id(record_id)
    if borrow_record:
        mold_model = borrow_record.get('mold_model', '')
        mold_position = borrow_record.get('mold_position', '')
        quantity = borrow_record.get('quantity', 1)
        status = borrow_record.get('status', '')
        borrower = borrow_record.get('borrower', '')
        borrow_date = borrow_record.get('borrow_date', '')
        # 如果是借出状态，需要归还刀模
        if status == '借出':
            _update_die_mold_borrow_status(mold_model, mold_position, quantity, is_return=True, borrower=borrower, borrow_date=borrow_date)

    delete_borrow_record(record_id)
    invalidate_inventory_cache()
    return jsonify({"success": True, "message": "删除成功"})


@app.route('/api/borrow/current')
def api_get_current_borrow():
    """获取所有刀模当前借出状态"""
    summary = get_borrow_summary_by_model()
    return jsonify({"success": True, "summary": summary})


@app.route('/api/borrow/export', methods=['GET'])
def api_export_borrow_records():
    """导出击用记录Excel"""
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    status = request.args.get('status')
    borrower = request.args.get('borrower')

    records = get_borrow_records(start_date, end_date, status, borrower)

    if not records:
        return jsonify({"success": False, "error": "没有可导出的记录"}), 400

    # 创建Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "领用记录"

    # 表头
    headers = ["ID", "刀模型号", "部位", "位置", "领用人", "领用日期", "数量", "归还日期", "归还人", "状态", "备注", "登记时间"]
    ws.append(headers)

    # 数据
    for r in records:
        ws.append([
            r['id'],
            r['mold_model'],
            r['mold_type'] or '',
            r['mold_position'] or '',
            r['borrower'],
            r['borrow_date'],
            r['quantity'],
            r['return_date'] or '',
            r['returner'] or '',
            r['status'],
            r['note'] or '',
            r['created_at']
        ])

    # 调整列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 30)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = BASE_DIR / f"刀模领用记录_{timestamp}.xlsx"
    wb.save(output_path)

    return send_file(output_path, as_attachment=True)


# ── API: 导出 ───────────────────────────────────────────
@app.route('/api/export')
def api_export():
    """导出库存 Excel（含借出状态标注）"""
    # 获取完整库存（含借出状态）
    records = load_die_molds()
    borrow_summary = get_borrow_summary_by_model_position()

    # 关联借出信息
    export_records = []
    for rec in records:
        model = rec.get('model', '')
        position = rec.get('position', '')
        borrow_info = borrow_summary.get((model, position), {})

        export_records.append({
            "位置": position,
            "刀模型号": model,
            "刀模类型": rec.get('mold_type', ''),
            "总数量": rec.get('total_quantity', 0),
            "已借出": rec.get('borrowed', 0),
            "可用库存": rec.get('available', 0),
            "借出状态": "借出中" if borrow_info else "在库",
            "借出人": borrow_info.get('borrower', ''),
            "借出日期": borrow_info.get('borrow_date', ''),
            "备注": rec.get('note', ''),
        })

    timestamp = datetime.now().strftime("%Y%m%d")
    output_path = BASE_DIR / f"刀模库存_{timestamp}.xlsx"
    export_to_excel(export_records, output_path)
    return send_file(output_path, as_attachment=True)


@app.route('/api/report/factories')
def api_report_factories():
    """获取所有打样工厂列表"""
    COL_FACTORY = 21  # 打样工厂列索引
    all_factories = set()
    current_year = datetime.now().year
    target_years = [current_year, current_year - 1]

    files = sorted(BASE_DIR.glob("生产工艺单*.xlsx"))
    for fpath in files:
        fname = fpath.name
        year_match = re.search(r'JF(\d{2})', fname)
        if not year_match:
            continue
        file_year_short = int(year_match.group(1))
        file_year = 2000 + file_year_short
        if file_year not in target_years:
            continue

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            if "Sheet1" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Sheet1"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                factory_val = row[COL_FACTORY] if COL_FACTORY < len(row) else None
                factory_name = str(factory_val).strip() if factory_val else ''
                if factory_name and factory_name != 'None':
                    all_factories.add(factory_name)

            wb.close()
        except Exception as e:
            print(f"  读取工艺单失败：{fpath.name} - {e}")
            continue

    return jsonify({"factories": sorted(all_factories)})


@app.route('/api/report/all')
def api_report_all():
    """获取缺失刀模报告 - 支持日期范围和打样工厂筛选"""
    # 需求日期列索引
    COL_REQUIRE_DATE = 17
    # 工艺单列索引
    COL_EYEBROW_TOP = 2  # 上眉
    COL_EYEBROW_BOT = 3  # 下眉
    COL_FRONT = 4  # 前片
    COL_SIDE = 5  # 边片
    COL_BACK = 6  # 后片
    COL_FACTORY = 21  # 打样工厂

    # 获取筛选参数
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    factory_filter = request.args.getlist('factory')  # 支持多选

    # 解析日期
    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    # 获取当前年份
    current_year = datetime.now().year  # 2026

    nonv_refs: dict[str, str] = {}  # ref -> type (前片/边片/后片)
    count_v: dict[str, int] = defaultdict(int)
    file_count = 0
    row_count = 0
    all_factories: set[str] = set()

    # 需要匹配的年份列表：当年和上一年
    target_years = [current_year, current_year - 1]  # [2026, 2025]

    # 遍历所有工艺单文件
    files = sorted(BASE_DIR.glob("生产工艺单*.xlsx"))

    for fpath in files:
        fname = fpath.name  # 如 "生产工艺单JF26.xlsx" 或 "生产工艺单JF25.xlsx"

        # 从文件名提取年份：JF26 -> 2026, JF25 -> 2025
        year_match = re.search(r'JF(\d{2})', fname)
        if not year_match:
            continue
        file_year_short = int(year_match.group(1))  # 26, 25
        file_year = 2000 + file_year_short  # 2026, 2025

        # 只处理当年和上一年的文件
        if file_year not in target_years:
            continue

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            if "Sheet1" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Sheet1"]

            n = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                # 获取打样工厂
                factory_val = row[COL_FACTORY] if COL_FACTORY < len(row) else None
                factory_name = str(factory_val).strip() if factory_val else ''
                if factory_name and factory_name != 'None':
                    all_factories.add(factory_name)

                # 打样工厂筛选（多选）
                if factory_filter and factory_name not in factory_filter:
                    continue

                # 解析需求日期
                req_date_val = row[COL_REQUIRE_DATE] if COL_REQUIRE_DATE < len(row) else None
                req_date = None
                if req_date_val:
                    try:
                        if isinstance(req_date_val, datetime):
                            req_date = req_date_val.date()
                        elif isinstance(req_date_val, str):
                            req_date = datetime.strptime(req_date_val.split()[0], "%Y-%m-%d").date()
                    except:
                        req_date = None

                # 如果有上一年文件，检查需求日期是否是当年
                if file_year == current_year - 1:
                    if req_date:
                        req_year = req_date.year
                        # 如果需求日期不是当年，跳过
                        if req_year != current_year:
                            continue
                    else:
                        continue

                # 日期范围筛选
                if req_date:
                    if start_date and req_date < start_date:
                        continue
                    if end_date and req_date > end_date:
                        continue

                # 收集非V型刀模引用（前/边/后片），带类型信息
                type_names = {COL_FRONT: '前片', COL_SIDE: '边片', COL_BACK: '后片'}
                for ci, type_name in type_names.items():
                    val = row[ci] if ci < len(row) else None
                    if val and isinstance(val, str):
                        key = normalize_process_ref(val)
                        if key:
                            nonv_refs[key] = type_name

                # 收集V型刀模引用（上眉/下眉）
                row_v_keys: set[str] = set()
                for ci in (COL_EYEBROW_TOP, COL_EYEBROW_BOT):
                    val = row[ci] if ci < len(row) else None
                    if val and isinstance(val, str) and val.strip():
                        row_v_keys.add(val.strip())
                for key in row_v_keys:
                    count_v[key] += 1

                n += 1

            print(f"  读取工艺单：{fname} ... {n:,} 行 (过滤后)")
            file_count += 1
            row_count += n
            wb.close()
        except Exception as e:
            print(f"  读取工艺单失败：{fname} - {e}")
            continue

    # 读取库存
    records = ml_load_die_molds(DIE_MOLD_FILE, DIE_MOLD_SHEET)

    # 收集工艺单中的所有引用
    all_refs = set(nonv_refs.keys()) | set(count_v.keys())

    # 构建库存索引（与刀模查询模块一致）
    inventory_models = {rec['model'] for rec in records}
    records_dict = {rec['model']: rec for rec in records}
    v_records = [rec for rec in records if rec['model'].startswith('V')]

    # 提取关键字的辅助函数
    def get_keyword_required(type_name):
        """从类型名称提取关键字"""
        if "前" in type_name:
            return "前"
        elif "边" in type_name:
            return "边"  # 边或侧都可以
        elif "后" in type_name:
            return "后"
        return None

    def note_has_keyword(rec, required):
        """检查备注或刀模型号是否包含关键字"""
        if not required:
            return True
        note = rec.get('note', '') or ''
        model = rec.get('model', '') or ''
        note_upper = note.upper()
        model_upper = model.upper()

        if required == "边":
            if "边" in note_upper or "侧" in note_upper or "边" in model_upper or "侧" in model_upper:
                return True
        else:
            if required in note_upper or required in model_upper:
                return True

        # 兜底：刀型号号无前/边/后关键字、数量>=3、备注无关键字 → 通用刀模
        model_has_dir = any(k in model_upper for k in ['前', '边', '后', '侧'])
        note_has_dir = any(k in note_upper for k in ['前', '边', '后', '侧'])
        total_qty = rec.get('total_quantity', 0) or rec.get('quantity', 0)
        if not model_has_dir and total_qty >= 3 and not note_has_dir:
            return True
        return False

    def match_v_mold(ref, v_records):
        """V型刀模精准匹配（与_match_v_mold_fast一致）"""
        ref = ref.strip()
        ref_seg_count = count_model_segments(ref)
        matched = []

        for rec in v_records:
            inv_model = rec['model']
            inv_seg_count = count_model_segments(inv_model)

            # 段数必须相同
            if inv_seg_count != ref_seg_count:
                continue

            # 1. 直接匹配
            if ref == inv_model:
                matched.append(rec)
                continue

            # 2. 包含匹配（双向，段数已相同）
            if ref in inv_model or inv_model in ref:
                matched.append(rec)
                continue

            # 3. 上下组合型（工艺单引用含"上下"）
            if "上下" in ref:
                base_ref = ref.replace("上下", "").replace("V", "")
                base_inv = inv_model.replace("上下", "").replace("V", "")
                if base_ref == base_inv:
                    matched.append(rec)
                    continue

            # 3.5 处理库存刀模是合并型写法（如 "V237舌上、下"）
            if "上、下" in inv_model or "上下" in inv_model:
                base = inv_model.replace("上、下", "").replace("上下", "")
                if ref.startswith(base) and ("上" in ref or "下" in ref):
                    matched.append(rec)
                    continue

            # 4. 舌字变体
            ref_no_tongue = ref.replace('舌', '')
            inv_no_tongue = inv_model.replace('舌', '')
            if ref_no_tongue == inv_no_tongue:
                matched.append(rec)
                continue
            # 上/下匹配
            if ref_no_tongue.endswith('上') and inv_no_tongue == ref_no_tongue[:-1] + '舌上':
                matched.append(rec)
                continue
            if ref_no_tongue.endswith('下') and inv_no_tongue == ref_no_tongue[:-1] + '舌下':
                matched.append(rec)
                continue

            # 5. 舌上/舌下 匹配（更严格的匹配）
            inv_norm = inv_model.replace('、', '').replace(',', '').replace('，', '')
            ref_norm = ref.replace('、', '').replace(',', '').replace('，', '')

            v_match = re.match(r'V(\d+)', ref_norm)
            inv_v_match = re.match(r'V(\d+)', inv_norm)
            if v_match and inv_v_match and v_match.group(1) == inv_v_match.group(1):
                ref_part = ref_norm.replace(v_match.group(0), '')
                inv_part = inv_norm.replace(inv_v_match.group(0), '')

                # 提取舌前部分
                if '舌' in ref_part:
                    idx = ref_part.index('舌')
                    ref_before_tongue = ref_part[:idx]
                else:
                    m = re.search(r'[上下]', ref_part)
                    ref_before_tongue = ref_part[:m.start()] if m else ''

                if '舌' in inv_part:
                    idx = inv_part.index('舌')
                    inv_before_tongue = inv_part[:idx]
                else:
                    m = re.search(r'[上下]', inv_part)
                    inv_before_tongue = inv_part[:m.start()] if m else ''

                # 舌前整体必须相同
                if ref_before_tongue != inv_before_tongue:
                    continue

                # 检查部位
                ref_has_upper = '上' in ref_part
                ref_has_lower = '下' in ref_part
                inv_has_upper = '上' in inv_part
                inv_has_lower = '下' in inv_part

                if ref_has_upper and not ref_has_lower and inv_has_upper:
                    matched.append(rec)
                elif ref_has_lower and not ref_has_upper and inv_has_lower:
                    matched.append(rec)
                elif ref_has_upper and ref_has_lower and inv_has_upper and inv_has_lower:
                    matched.append(rec)

        return matched

    def match_nonv_mold(ref, type_name, records_dict, inventory_models):
        """非V型刀模精准匹配（与_match_nonv_mold_by_prefix一致）"""
        keyword_required = get_keyword_required(type_name)
        ref_seg_count = count_model_segments(ref)
        ref_keys = get_die_mold_keys(ref)
        matched = []

        for inv_model in inventory_models:
            rec = records_dict.get(inv_model)
            if not rec:
                continue

            inv_seg_count = count_model_segments(inv_model)

            # 段数必须相同
            if inv_seg_count != ref_seg_count:
                continue

            inv_keys = get_die_mold_keys(inv_model)

            # 键集合交集
            if ref_keys & inv_keys:
                if note_has_keyword(rec, keyword_required):
                    matched.append(rec)

        return matched

    # 找出缺失
    missing = []
    for ref in all_refs:
        is_v = ref in count_v
        found = False

        if is_v:
            # V型匹配
            matched = match_v_mold(ref, v_records)
            if matched:
                found = True
        else:
            # 非V型匹配
            type_name = nonv_refs.get(ref, '')
            matched = match_nonv_mold(ref, type_name, records_dict, inventory_models)
            if matched:
                found = True

        if not found:
            mold_type = "V型" if is_v else "非V型"
            # 确定缺失类型
            if ref in count_v or any('眉' in ref for ref in count_v):
                type_detail = "上眉" if '上' in ref or '眉' in ref else "下眉"
            elif ref in nonv_refs:
                type_detail = nonv_refs[ref]
            else:
                type_detail = mold_type

            missing.append({
                "model": ref,
                "type": type_detail,
                "quantity": 1
            })

    # 去重
    seen = set()
    unique_missing = []
    for m in missing:
        if m['model'] not in seen:
            seen.add(m['model'])
            unique_missing.append(m)

    return jsonify({
        "file_count": file_count,
        "row_count": row_count,
        "missing": unique_missing,
        "factories": sorted(all_factories),
        "summary": {
            "total_missing": len(unique_missing)
        }
    })


@app.route('/api/report/export/excel')
def api_report_export_excel():
    """导出缺失刀模报告为 Excel 文件"""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    factory_filter = request.args.getlist('factory')  # 支持多选

    # 复用 api_report_all 的逻辑获取缺失数据
    COL_REQUIRE_DATE = 17
    COL_EYEBROW_TOP = 2
    COL_EYEBROW_BOT = 3
    COL_FRONT = 4
    COL_SIDE = 5
    COL_BACK = 6
    COL_FACTORY = 21

    start_date = None
    end_date = None
    if start_date_str:
        try:
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    if end_date_str:
        try:
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    current_year = datetime.now().year
    nonv_refs: dict[str, str] = {}  # ref -> type (前片/边片/后片)
    count_v: dict[str, int] = defaultdict(int)
    target_years = [current_year, current_year - 1]

    files = sorted(BASE_DIR.glob("生产工艺单*.xlsx"))
    for fpath in files:
        fname = fpath.name
        year_match = re.search(r'JF(\d{2})', fname)
        if not year_match:
            continue
        file_year_short = int(year_match.group(1))
        file_year = 2000 + file_year_short

        if file_year not in target_years:
            continue

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
            if "Sheet1" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Sheet1"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                factory_val = row[COL_FACTORY] if COL_FACTORY < len(row) else None
                factory_name = str(factory_val).strip() if factory_val else ''
                if factory_filter and factory_name not in factory_filter:
                    continue

                req_date_val = row[COL_REQUIRE_DATE] if COL_REQUIRE_DATE < len(row) else None
                req_date = None
                if req_date_val:
                    try:
                        if isinstance(req_date_val, datetime):
                            req_date = req_date_val.date()
                        elif isinstance(req_date_val, str):
                            req_date = datetime.strptime(req_date_val.split()[0], "%Y-%m-%d").date()
                    except:
                        req_date = None

                if file_year == current_year - 1:
                    if req_date:
                        req_year = req_date.year
                        if req_year != current_year:
                            continue
                    else:
                        continue

                if req_date:
                    if start_date and req_date < start_date:
                        continue
                    if end_date and req_date > end_date:
                        continue

                # 收集非V型刀模引用（前/边/后片），带类型信息
                type_names = {COL_FRONT: '前片', COL_SIDE: '边片', COL_BACK: '后片'}
                for ci, type_name in type_names.items():
                    val = row[ci] if ci < len(row) else None
                    if val and isinstance(val, str):
                        key = normalize_process_ref(val)
                        if key:
                            nonv_refs[key] = type_name

                row_v_keys: set[str] = set()
                for ci in (COL_EYEBROW_TOP, COL_EYEBROW_BOT):
                    val = row[ci] if ci < len(row) else None
                    if val and isinstance(val, str) and val.strip():
                        row_v_keys.add(val.strip())
                for key in row_v_keys:
                    count_v[key] += 1

            wb.close()
        except Exception as e:
            print(f"  读取工艺单失败：{fpath.name} - {e}")
            continue

    records = ml_load_die_molds(DIE_MOLD_FILE, DIE_MOLD_SHEET)
    all_refs = set(nonv_refs.keys()) | set(count_v.keys())

    # 构建库存索引
    inventory_models = {rec['model'] for rec in records}
    records_dict = {rec['model']: rec for rec in records}
    v_records = [rec for rec in records if rec['model'].startswith('V')]

    # 提取关键字的辅助函数
    def get_keyword_required(type_name):
        if "前" in type_name:
            return "前"
        elif "边" in type_name:
            return "边"
        elif "后" in type_name:
            return "后"
        return None

    def note_has_keyword(rec, required):
        if not required:
            return True
        note = rec.get('note', '') or ''
        model = rec.get('model', '') or ''
        note_upper = note.upper()
        model_upper = model.upper()

        if required == "边":
            if "边" in note_upper or "侧" in note_upper or "边" in model_upper or "侧" in model_upper:
                return True
        else:
            if required in note_upper or required in model_upper:
                return True

        model_has_dir = any(k in model_upper for k in ['前', '边', '后', '侧'])
        note_has_dir = any(k in note_upper for k in ['前', '边', '后', '侧'])
        total_qty = rec.get('total_quantity', 0) or rec.get('quantity', 0)
        if not model_has_dir and total_qty >= 3 and not note_has_dir:
            return True
        return False

    def match_v_mold(ref, v_records):
        """V型刀模精准匹配"""
        ref = ref.strip()
        ref_seg_count = count_model_segments(ref)
        matched = []

        for rec in v_records:
            inv_model = rec['model']
            inv_seg_count = count_model_segments(inv_model)
            if inv_seg_count != ref_seg_count:
                continue

            # 1. 直接匹配
            if ref == inv_model:
                matched.append(rec)
                continue

            # 2. 包含匹配
            if ref in inv_model or inv_model in ref:
                matched.append(rec)
                continue

            # 3. 上下组合型（工艺单引用含"上下"）
            if "上下" in ref:
                base_ref = ref.replace("上下", "").replace("V", "")
                base_inv = inv_model.replace("上下", "").replace("V", "")
                if base_ref == base_inv:
                    matched.append(rec)
                    continue

            # 3.5 处理库存刀模是合并型写法（如 "V237舌上、下"）
            if "上、下" in inv_model or "上下" in inv_model:
                base = inv_model.replace("上、下", "").replace("上下", "")
                if ref.startswith(base) and ("上" in ref or "下" in ref):
                    matched.append(rec)
                    continue

            # 4. 舌字变体
            ref_no_tongue = ref.replace('舌', '')
            inv_no_tongue = inv_model.replace('舌', '')
            if ref_no_tongue == inv_no_tongue:
                matched.append(rec)
                continue
            if ref_no_tongue.endswith('上') and inv_no_tongue == ref_no_tongue[:-1] + '舌上':
                matched.append(rec)
                continue
            if ref_no_tongue.endswith('下') and inv_no_tongue == ref_no_tongue[:-1] + '舌下':
                matched.append(rec)
                continue

            # 5. 舌上/舌下 匹配
            inv_norm = inv_model.replace('、', '').replace(',', '').replace('，', '')
            ref_norm = ref.replace('、', '').replace(',', '').replace('，', '')

            v_match = re.match(r'V(\d+)', ref_norm)
            inv_v_match = re.match(r'V(\d+)', inv_norm)
            if v_match and inv_v_match and v_match.group(1) == inv_v_match.group(1):
                ref_part = ref_norm.replace(v_match.group(0), '')
                inv_part = inv_norm.replace(inv_v_match.group(0), '')

                if '舌' in ref_part:
                    idx = ref_part.index('舌')
                    ref_before_tongue = ref_part[:idx]
                else:
                    m = re.search(r'[上下]', ref_part)
                    ref_before_tongue = ref_part[:m.start()] if m else ''

                if '舌' in inv_part:
                    idx = inv_part.index('舌')
                    inv_before_tongue = inv_part[:idx]
                else:
                    m = re.search(r'[上下]', inv_part)
                    inv_before_tongue = inv_part[:m.start()] if m else ''

                if ref_before_tongue != inv_before_tongue:
                    continue

                ref_has_upper = '上' in ref_part
                ref_has_lower = '下' in ref_part
                inv_has_upper = '上' in inv_part
                inv_has_lower = '下' in inv_part

                if ref_has_upper and not ref_has_lower and inv_has_upper:
                    matched.append(rec)
                elif ref_has_lower and not ref_has_upper and inv_has_lower:
                    matched.append(rec)
                elif ref_has_upper and ref_has_lower and inv_has_upper and inv_has_lower:
                    matched.append(rec)

        return matched

    def match_nonv_mold(ref, type_name, records_dict, inventory_models):
        """非V型刀模精准匹配"""
        keyword_required = get_keyword_required(type_name)
        ref_seg_count = count_model_segments(ref)
        ref_keys = get_die_mold_keys(ref)
        matched = []

        for inv_model in inventory_models:
            rec = records_dict.get(inv_model)
            if not rec:
                continue
            inv_seg_count = count_model_segments(inv_model)
            if inv_seg_count != ref_seg_count:
                continue
            inv_keys = get_die_mold_keys(inv_model)
            if ref_keys & inv_keys:
                if note_has_keyword(rec, keyword_required):
                    matched.append(rec)
        return matched

    missing = []
    for ref in all_refs:
        is_v = ref in count_v
        found = False

        if is_v:
            matched = match_v_mold(ref, v_records)
            if matched:
                found = True
        else:
            type_name = nonv_refs.get(ref, '')
            matched = match_nonv_mold(ref, type_name, records_dict, inventory_models)
            if matched:
                found = True

        if not found:
            mold_type = "V型" if is_v else "非V型"
            if ref in count_v or any('眉' in ref for ref in count_v):
                type_detail = "上眉" if '上' in ref or '眉' in ref else "下眉"
            elif ref in nonv_refs:
                type_detail = nonv_refs[ref]
            else:
                type_detail = mold_type

            missing.append({
                "model": ref,
                "type": type_detail,
                "quantity": 1
            })

    # 去重
    seen = set()
    unique_missing = []
    for m in missing:
        if m['model'] not in seen:
            seen.add(m['model'])
            unique_missing.append(m)

    # 创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "缺失刀模采购清单"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4472C4")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 写入表头
    headers = ["序号", "刀模型号", "类型", "建议数量"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 写入数据
    for row_idx, m in enumerate(unique_missing, 2):
        ws.cell(row=row_idx, column=1, value=row_idx - 1).border = thin_border
        ws.cell(row=row_idx, column=2, value=m["model"]).border = thin_border
        ws.cell(row=row_idx, column=3, value=m["type"]).border = thin_border
        ws.cell(row=row_idx, column=4, value=m["quantity"]).border = thin_border

    # 设置列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 12

    # 保存文件
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_path = BASE_DIR / f"刀模采购清单_{timestamp}.xlsx"
    wb.save(output_path)
    wb.close()

    return send_file(output_path, as_attachment=True)


# ── API: 按分发时间/裁剪计划分析缺料 ───────────────────────────────────────────
def _parse_date(val):
    """解析日期值，返回 datetime.date 或 None"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        val = val.strip()
        if not val:
            return None
        # 尝试多种日期格式
        for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"):
            try:
                return datetime.strptime(val, fmt).date()
            except ValueError:
                continue
    return None


def _get_last_workday():
    """获取上一个工作日（周一到周五）"""
    today = date.today()
    weekday = today.weekday()
    if weekday == 0:  # 周一
        return today - timedelta(days=3)  # 上周五
    elif weekday == 6:  # 周日
        return today - timedelta(days=2)  # 上周五
    else:  # 周二到周六
        return today - timedelta(days=1)


def _analyze_mould_shortage(date_type, start_date=None, end_date=None):
    """分析刀模缺料情况

    策略：
    1. 从大货生产计划跟踪表获取日期筛选后的计划跟踪号及其日期信息
    2. 从生产工艺单获取这些计划跟踪号的完整刀模数据（上眉、下眉、前片、边片、后片）
    3. 使用现有匹配算法检查库存

    Args:
        date_type: 'distribution' 或 'cutting_plan'
        start_date: 分析开始日期（inclusive）
        end_date: 分析结束日期（inclusive）

    Returns:
        (Excel workbook, results list)
    """
    from datetime import date, timedelta

    tracking_file = BASE_DIR / "大货生产计划跟踪表.xlsx"
    if not tracking_file.exists():
        print(f"文件不存在: {tracking_file}")
        return None, []

    today = date.today()
    last_workday = _get_last_workday()

    # 确定日期范围（仅裁剪计划用日期范围，distribution用分发时间非空+裁剪实际为空）
    if date_type == 'distribution':
        # distribution: 分发时间非空 且 裁剪实际为空，不设日期范围
        pass
    else:  # cutting_plan
        if start_date is None:
            start_date = today
        if end_date is None:
            end_date = today + timedelta(days=365)

    # ========== 第1步：从大货生产计划跟踪表获取日期筛选后的计划跟踪号 ==========
    print(f"读取文件: {tracking_file}")
    wb_track = openpyxl.load_workbook(tracking_file, data_only=True, read_only=True)
    ws_track = wb_track.active

    # 获取表头并建立列名到索引的映射
    headers_track = [str(cell.value).strip() if cell.value else f'_col_{i}' for i, cell in enumerate(next(ws_track.iter_rows(min_row=1, max_row=1)))]
    col_map_track = {h: i for i, h in enumerate(headers_track)}

    col_track_no = col_map_track.get('生产单号', 2)
    col_dist_time = col_map_track.get('分发时间', 12)
    col_cut_plan = col_map_track.get('裁剪计划', 14)
    col_cut_actual = col_map_track.get('裁剪实际', None)  # 裁剪实际列，可能不存在
    col_workshop = col_map_track.get('生产车间', None)  # 生产车间列

    # 收集符合条件的计划跟踪号及其日期信息
    track_info = {}  # {track_no: {dist_time, cut_plan}}
    for row in ws_track.iter_rows(min_row=2, values_only=True):
        if not row or len(row) <= col_track_no:
            continue

        track_no = str(row[col_track_no]).strip() if row[col_track_no] else ''
        if not track_no:
            continue

        # 获取日期
        dist_time = row[col_dist_time] if col_dist_time < len(row) else None
        cut_plan = row[col_cut_plan] if col_cut_plan < len(row) else None
        cut_actual = row[col_cut_actual].strip() if col_cut_actual is not None and col_cut_actual < len(row) and row[col_cut_actual] else ''

        # 排除泰州七车间和泰州生产部
        workshop = str(row[col_workshop]).strip() if col_workshop is not None and col_workshop < len(row) and row[col_workshop] else ''
        if workshop in ('泰州七车间', '泰州生产部'):
            continue

        if date_type == 'distribution':
            # 筛选条件：分发时间有值 且 裁剪实际为空
            dist_time_str = str(dist_time).strip() if dist_time else ''
            if not dist_time_str or cut_actual:
                continue
        else:
            # 裁剪计划：按日期范围过滤
            date_col_idx = col_cut_plan
            date_val = row[date_col_idx] if date_col_idx < len(row) else None
            order_date = _parse_date(date_val)

            if order_date is None:
                continue

            # 日期过滤
            if not (start_date <= order_date <= end_date):
                continue

        if track_no not in track_info:
            track_info[track_no] = {
                'distribution_time': str(dist_time)[:10] if dist_time else '',
                'cutting_plan': str(cut_plan)[:10] if cut_plan else ''
            }

    wb_track.close()
    print(f"  符合日期条件的计划跟踪号: {len(track_info)}")

    if not track_info:
        return None, []

    # ========== 第2步：从生产工艺单获取刀模数据 ==========
    # 工艺单列索引
    _COL_TRACK_NO = 0
    _COL_EYEBROW_TOP = 2
    _COL_EYEBROW_BOT = 3
    _COL_FRONT = 4
    _COL_SIDE = 5
    _COL_BACK = 6

    # 收集工艺单中的刀模需求
    process_molds = []  # [{track_no, position, model, dist_time, cut_plan}]
    files = sorted(BASE_DIR.glob("生产工艺单*.xlsx"))

    for fpath in files:
        try:
            wb = openpyxl.load_workbook(fpath, data_only=True, read_only=True)
            if "Sheet1" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Sheet1"]

            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) <= _COL_TRACK_NO:
                    continue

                track_no = str(row[_COL_TRACK_NO]).strip() if row[_COL_TRACK_NO] else ''
                if not track_no:
                    continue

                # 检查是否在筛选的计划跟踪号列表中
                if track_no not in track_info:
                    continue

                info = track_info[track_no]
                dist_time = info['distribution_time']
                cut_plan = info['cutting_plan']

                # 处理上眉 (V型)
                if len(row) > _COL_EYEBROW_TOP and row[_COL_EYEBROW_TOP] and isinstance(row[_COL_EYEBROW_TOP], str):
                    mold_model = row[_COL_EYEBROW_TOP].strip()
                    if mold_model:
                        process_molds.append({
                            'track_no': track_no,
                            'position': '上眉',
                            'model': mold_model,
                            'mold_type': 'V型',
                            'distribution_time': dist_time,
                            'cutting_plan': cut_plan
                        })

                # 处理下眉 (V型)
                if len(row) > _COL_EYEBROW_BOT and row[_COL_EYEBROW_BOT] and isinstance(row[_COL_EYEBROW_BOT], str):
                    mold_model = row[_COL_EYEBROW_BOT].strip()
                    if mold_model:
                        process_molds.append({
                            'track_no': track_no,
                            'position': '下眉',
                            'model': mold_model,
                            'mold_type': 'V型',
                            'distribution_time': dist_time,
                            'cutting_plan': cut_plan
                        })

                # 处理前片 (非V型)
                if len(row) > _COL_FRONT and row[_COL_FRONT] and isinstance(row[_COL_FRONT], str):
                    mold_model = row[_COL_FRONT].strip()
                    if mold_model:
                        process_molds.append({
                            'track_no': track_no,
                            'position': '前片',
                            'model': mold_model,
                            'mold_type': '非V型',
                            'distribution_time': dist_time,
                            'cutting_plan': cut_plan
                        })

                # 处理边片 (非V型)
                if len(row) > _COL_SIDE and row[_COL_SIDE] and isinstance(row[_COL_SIDE], str):
                    mold_model = row[_COL_SIDE].strip()
                    if mold_model:
                        process_molds.append({
                            'track_no': track_no,
                            'position': '边片',
                            'model': mold_model,
                            'mold_type': '非V型',
                            'distribution_time': dist_time,
                            'cutting_plan': cut_plan
                        })

                # 处理后片 (非V型)
                if len(row) > _COL_BACK and row[_COL_BACK] and isinstance(row[_COL_BACK], str):
                    mold_model = row[_COL_BACK].strip()
                    if mold_model:
                        process_molds.append({
                            'track_no': track_no,
                            'position': '后片',
                            'model': mold_model,
                            'mold_type': '非V型',
                            'distribution_time': dist_time,
                            'cutting_plan': cut_plan
                        })

            wb.close()
        except Exception as e:
            print(f"  读取工艺单失败: {fpath.name} - {e}")
            continue

    print(f"  工艺单中匹配到的刀模需求: {len(process_molds)}")

    # ========== 第3步：匹配刀模 ==========
    cache = get_inventory_cache()
    inventory_models = cache['models_index']
    v_models = cache['v_models']
    nonv_prefix_index = cache['nonv_prefix_index']
    nonv_full_index = cache['nonv_full_index']

    results = []
    for pm in process_molds:
        mold_infos = check_mold_in_stock_fast(
            pm['model'], pm['position'], pm['mold_type'],
            inventory_models, v_models, nonv_prefix_index, nonv_full_index
        )
        for mi in mold_infos:
            results.append({
                "track_no": pm['track_no'],
                "distribution_time": pm['distribution_time'],
                "cutting_plan": pm['cutting_plan'],
                "position": pm['position'],
                "required_model": pm['model'],
                "match_status": "有刀模" if mi.get('in_stock') else "缺失",
                "inventory_model": mi.get('inventory_model') or '',
                "mold_position": mi.get('position') or '',
                "note": mi.get('note') or '',
            })

    print(f"  匹配结果: {len(results)} 条")

    # ========== 第4步：生成Excel ==========
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "缺料分析"

    headers = ["计划跟踪号", "分发时间", "裁剪计划", "部位", "需求型号", "匹配状态", "匹配刀模", "刀模部位", "备注"]
    ws.append(headers)

    from openpyxl.styles import Font, PatternFill, Alignment
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type="solid", fgColor="D9E1F2")
    header_alignment = Alignment(horizontal='center', vertical='center')

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    red_fill = PatternFill(fill_type="solid", fgColor="FFCCCC")
    red_font = Font(color="FF0000", bold=True)
    green_font = Font(color="00B050")

    for row_data in results:
        row = [
            row_data['track_no'],
            row_data['distribution_time'],
            row_data['cutting_plan'],
            row_data['position'],
            row_data['required_model'],
            row_data['match_status'],
            row_data['inventory_model'],
            row_data['mold_position'],
            row_data['note'],
        ]
        ws.append(row)

        row_idx = ws.max_row
        in_stock = row_data['match_status'] == '有刀模'

        if not in_stock:
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = red_fill
            ws.cell(row=row_idx, column=6).font = red_font
        else:
            ws.cell(row=row_idx, column=6).font = green_font

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 10
    ws.column_dimensions["G"].width = 22
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 15

    return wb, results


@app.route('/api/analyze/by_distribution')
def api_analyze_by_distribution():
    """按分发时间分析缺料：从上一个工作日到今天的订单"""
    try:
        wb, results = _analyze_mould_shortage('distribution')
        if wb is None and not results:
            return jsonify({"error": "没有数据"}), 404

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = BASE_DIR / f"缺料分析_按分发时间_{timestamp}.xlsx"
        wb.save(output_path)

        print(f"生成文件: {output_path}, 记录数: {len(results)}")
        return send_file(output_path, as_attachment=True)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/analyze/by_cutting_plan')
def api_analyze_by_cutting_plan():
    """按裁剪计划分析缺料：今天之后的订单"""
    try:
        wb, results = _analyze_mould_shortage('cutting_plan')
        if wb is None and not results:
            return jsonify({"error": "没有数据"}), 404

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = BASE_DIR / f"缺料分析_按裁剪计划_{timestamp}.xlsx"
        wb.save(output_path)

        print(f"生成文件: {output_path}, 记录数: {len(results)}")
        return send_file(output_path, as_attachment=True)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": str(e)}), 500


@app.route('/api/analyze/files')
def api_analyze_files():
    """获取已生成的缺料分析文件列表"""
    import os

    files = []
    for f in BASE_DIR.glob("缺料分析_*.xlsx"):
        stat = f.stat()
        files.append({
            "name": f.name,
            "size": stat.st_size,
            "created": datetime.fromtimestamp(stat.st_ctime).strftime("%Y-%m-%d %H:%M:%S"),
            "path": f"/api/analyze/download/{f.name}"
        })

    # 按创建时间倒序排列
    files.sort(key=lambda x: x["created"], reverse=True)
    return jsonify(files)


@app.route('/api/analyze/download/<filename>')
def api_analyze_download(filename):
    """下载缺料分析文件"""
    from werkzeug.utils import secure_filename
    # 安全检查：只允许下载缺料分析文件
    if not filename.startswith("缺料分析_") or not filename.endswith(".xlsx"):
        return jsonify({"error": "无效的文件名"}), 400

    output_path = BASE_DIR / filename
    if not output_path.exists():
        return jsonify({"error": "文件不存在"}), 404

    return send_file(output_path, as_attachment=True)


@app.route('/api/analyze/delete/<filename>', methods=['DELETE'])
def api_analyze_delete(filename):
    """删除缺料分析文件"""
    # 安全检查：只允许删除缺料分析文件
    if not filename.startswith("缺料分析_") or not filename.endswith(".xlsx"):
        return jsonify({"error": "无效的文件名"}), 400

    output_path = BASE_DIR / filename
    if not output_path.exists():
        return jsonify({"error": "文件不存在"}), 404

    try:
        output_path.unlink()
        return jsonify({"success": True, "message": f"已删除 {filename}"})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ==================== 数据导入模块 ====================

@app.route('/import')
def import_page():
    """数据导入管理页面"""
    if 'username' not in session:
        return redirect('/login')
    return render_template('import.html')


@app.route('/api/import/verify_password', methods=['POST'])
def api_import_verify_password():
    """验证数据导入密码"""
    if 'username' not in session:
        return jsonify({"success": False, "error": "未登录"}), 401

    data = request.get_json()
    password = data.get('password', '')

    from config import IMPORT_PASSWORD
    if password == IMPORT_PASSWORD:
        return jsonify({"success": True})
    else:
        return jsonify({"success": False, "error": "密码错误"})


@app.route('/api/import/change_password', methods=['POST'])
def api_import_change_password():
    """修改数据导入密码"""
    if 'username' not in session:
        return jsonify({"success": False, "error": "未登录"}), 401

    data = request.get_json()
    current_password = data.get('current_password', '')
    new_password = data.get('new_password', '')

    from config import IMPORT_PASSWORD
    if current_password != IMPORT_PASSWORD:
        return jsonify({"success": False, "error": "当前密码错误"})

    if len(new_password) < 4:
        return jsonify({"success": False, "error": "新密码长度至少4位"})

    # 更新 config.py 中的密码
    config_path = BASE_DIR / 'config.py'
    with open(config_path, 'r', encoding='utf-8') as f:
        content = f.read()

    import re
    new_content = re.sub(
        r"(^IMPORT_PASSWORD\s*=\s*)['\"](.+?)['\"]",
        r"\1'" + new_password + "'",
        content,
        flags=re.MULTILINE
    )

    with open(config_path, 'w', encoding='utf-8') as f:
        f.write(new_content)

    return jsonify({"success": True, "message": "密码修改成功"})


@app.route('/api/import/die_mold', methods=['POST'])
def api_import_die_mold():
    """导入刀模汇总表"""
    if 'username' not in session:
        return jsonify({"error": "未登录"}), 401

    if 'file' not in request.files:
        return jsonify({"error": "没有上传文件"}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "只支持 .xlsx 或 .xls 文件"}), 400

    mode = request.form.get('mode', 'full_replace')
    if mode not in ['full_replace', 'incremental']:
        mode = 'full_replace'

    # 保存临时文件
    import tempfile
    import os
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name

    try:
        from utils.die_mold_importer import import_die_mold
        print(f"[DEBUG api] tmp_path={tmp_path}, mode={mode}")
        result = import_die_mold(tmp_path, mode, session.get('username', 'unknown'))
        print(f"[DEBUG api] result={result}")

        # 清除库存缓存并立即重新加载
        invalidate_inventory_cache()
        # 预热缓存
        get_inventory_cache()

        return jsonify(result)
    finally:
        os.unlink(tmp_path)


@app.route('/api/import/process', methods=['POST'])
def api_import_process():
    """导入生产工艺单"""
    if 'username' not in session:
        return jsonify({"error": "未登录"}), 401

    if 'file' not in request.files:
        return jsonify({"error": "没有上传文件"}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "只支持 .xlsx 或 .xls 文件"}), 400

    mode = request.form.get('mode', 'full_replace')
    if mode not in ['full_replace', 'incremental']:
        mode = 'full_replace'

    # 保存临时文件（保留原始文件名）
    import tempfile
    import os
    original_filename = file.filename
    tmp_path = f"/tmp/{original_filename}"
    file.save(tmp_path)

    try:
        from utils.process_importer import import_process
        result = import_process(tmp_path, mode, session.get('username', 'unknown'))

        # 立即重建工艺单缓存
        from utils.process_cache import rebuild_cache
        rebuild_cache()

        return jsonify(result)
    finally:
        os.unlink(tmp_path)


@app.route('/api/import/production_track', methods=['POST'])
def api_import_production_track():
    """导入大货生产计划跟踪表"""
    if 'username' not in session:
        return jsonify({"error": "未登录"}), 401

    if 'file' not in request.files:
        return jsonify({"error": "没有上传文件"}), 400

    file = request.files['file']
    if not file.filename.endswith(('.xlsx', '.xls')):
        return jsonify({"error": "只支持 .xlsx 或 .xls 文件"}), 400

    mode = request.form.get('mode', 'full_replace')
    if mode not in ['full_replace', 'incremental']:
        mode = 'full_replace'

    # 保存临时文件
    import tempfile
    import os
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        file.save(tmp.name)
        tmp_path = tmp.name

    try:
        from utils.production_importer import import_production_tracking
        result = import_production_tracking(tmp_path, mode, session.get('username', 'unknown'))
        return jsonify(result)
    finally:
        os.unlink(tmp_path)


@app.route('/api/import/template/<import_type>', methods=['GET'])
def api_import_template(import_type):
    """下载导入模板"""
    if 'username' not in session:
        return jsonify({"error": "未登录"}), 401

    import io
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active

    if import_type == 'die_mold':
        ws.title = '刀模汇总表'
        headers = ['位置', '刀模型号', '刀模类型', '总数量', '已借出', '可用库存', '备注']
        ws.append(headers)
        # 示例数据
        ws.append(['539', 'A1-523-59-050', 'V型', 3, 1, 2, '不带牙剪'])
        ws.append(['306', 'B2-123-45-001', '非V型', 5, 0, 5, ''])
    elif import_type == 'process':
        ws.title = '生产工艺单'
        headers = ['计划跟踪号', '客户名称', '上眉', '下眉', '前片', '边片', '后片', '盔型', '帽型']
        ws.append(headers)
        ws.append(['JF2601001', '客户A', 'A1-523-59-050', 'A1-523-59-051', 'B2-123-45-001', 'B2-123-45-002', 'B2-123-45-003', '盔型A', '帽型X'])
    elif import_type == 'production_track':
        ws.title = '生产跟踪表'
        headers = ['计划跟踪号', '工厂', '产品名称', '数量', '切割方案', '分发时间']
        ws.append(headers)
        ws.append(['JF2601001', '工厂A', '产品X', 1000, '方案1', '2026-04-25'])
    else:
        return jsonify({"error": "无效的模板类型"}), 400

    # 设置列宽
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{import_type}_import_template.xlsx"
    return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=filename)


if __name__ == '__main__':
    print("=" * 50)
    print("刀模管理系统 Web 版")
    print("=" * 50)

    # 启动时自动检查工艺单缓存
    print("\n[启动检查] 工艺单缓存状态:")
    from utils.process_cache import rebuild_cache, get_cache_stats
    stats = get_cache_stats()
    print(f"  缓存记录数: {stats['count']:,}")
    print("  检查文件变化...")
    rebuild_cache()

    print("\n访问 http://localhost:5000")
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False)
