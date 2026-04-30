#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
后整理计划表数据处理脚本
功能：
1. 将未完成工作表中实际已完成的行移到已完成工作表
2. 将大货生产计划表中的新单号添加到未完成工作表
"""

import pandas as pd
from datetime import datetime
import os
import shutil
from openpyxl import load_workbook

import tkinter as tk
from tkinter import filedialog, messagebox, scrolledtext


def create_gui():
    """创建 GUI 界面"""
    def select_file(entry_widget):
        """打开文件选择对话框"""
        filename = filedialog.askopenfilename(
            title="选择文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if filename:
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, filename)

    def on_start():
        """开始处理按钮回调"""
        main_file = entry_main.get().strip()
        schedule_file = entry_schedule.get().strip()

        if not main_file:
            messagebox.showwarning("警告", "请选择或输入「大货生产计划表」文件路径")
            return
        if not schedule_file:
            messagebox.showwarning("警告", "请选择或输入「后整理计划表」文件路径")
            return

        if not os.path.exists(main_file):
            messagebox.showerror("错误", f"文件不存在：{main_file}")
            return
        if not os.path.exists(schedule_file):
            messagebox.showerror("错误", f"文件不存在：{schedule_file}")
            return

        # 禁用开始按钮
        btn_start.config(state=tk.DISABLED)
        # 清空日志
        text_log.delete(1.0, tk.END)

        class TextRedirector:
            def __init__(self, widget):
                self.widget = widget
            def write(self, text):
                self.widget.insert(tk.END, text)
                self.widget.see(tk.END)
                self.widget.update_idletasks()
            def flush(self):
                pass

        import sys
        old_stdout = sys.stdout

        try:
            # 重定向 print 输出到日志文本框
            sys.stdout = TextRedirector(text_log)

            process_data(main_file=main_file, schedule_file=schedule_file)

            sys.stdout = old_stdout
            messagebox.showinfo("完成", "处理成功！")
        except Exception as e:
            sys.stdout = old_stdout
            messagebox.showerror("错误", f"处理失败：{str(e)}")
        finally:
            btn_start.config(state=tk.NORMAL)

    # 创建主窗口
    root = tk.Tk()
    root.title("后整理计划表处理")
    root.geometry("700x500")
    root.resizable(True, True)

    # 大货生产计划表
    frame_main = tk.Frame(root, pady=10)
    frame_main.pack(fill=tk.X, padx=20)
    tk.Label(frame_main, text="大货生产计划表（参照表）:").pack(anchor=tk.W)
    frame_main_input = tk.Frame(frame_main)
    frame_main_input.pack(fill=tk.X, pady=5)
    entry_main = tk.Entry(frame_main_input)
    entry_main.insert(0, "大货生产计划表.xlsx")
    entry_main.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    tk.Button(frame_main_input, text="浏览...", command=lambda: select_file(entry_main)).pack(side=tk.RIGHT)

    # 后整理计划表
    frame_schedule = tk.Frame(root, pady=10)
    frame_schedule.pack(fill=tk.X, padx=20)
    tk.Label(frame_schedule, text="后整理计划表（操作表）:").pack(anchor=tk.W)
    frame_schedule_input = tk.Frame(frame_schedule)
    frame_schedule_input.pack(fill=tk.X, pady=5)
    entry_schedule = tk.Entry(frame_schedule_input)
    entry_schedule.insert(0, "后整理计划表.xlsx")
    entry_schedule.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
    tk.Button(frame_schedule_input, text="浏览...", command=lambda: select_file(entry_schedule)).pack(side=tk.RIGHT)

    # 日志区域
    tk.Label(root, text="处理日志:").pack(anchor=tk.W, padx=20, pady=(10, 5))
    text_log = scrolledtext.ScrolledText(root, height=8, width=80)
    text_log.pack(fill=tk.BOTH, expand=False, padx=20, pady=5)

    # 开始按钮
    btn_start = tk.Button(root, text="开始处理", command=on_start, bg="#4CAF50", fg="white", padx=20, pady=5)
    btn_start.pack(pady=15)

    root.mainloop()


def process_data(main_file='大货生产计划表.xlsx', schedule_file='后整理计划表.xlsx'):
    """处理后整理计划表数据"""
    # 备份原文件
    schedule_backup = schedule_file.replace('.xlsx', '_backup.xlsx')
    if os.path.exists(schedule_file):
        shutil.copy(schedule_file, schedule_backup)
        print(f"已备份原文件：{schedule_backup}")

    # 读取文件
    print("正在读取文件...")
    df_main = pd.read_excel(main_file)
    df_todo = pd.read_excel(schedule_file, sheet_name='未完成')
    df_done = pd.read_excel(schedule_file, sheet_name='已完成')

    today = pd.Timestamp.now().normalize()
    print(f"今天日期：{today}")

    # ============================================================
    # 步骤 0: 过滤未完成工作表，删除大货生产计划表中没有的单号
    # ============================================================
    print("\n=== 过滤未完成工作表 ===")

    df_main['生产单号'] = df_main['生产单号'].astype(str)
    df_todo['生产单号'] = df_todo['生产单号'].astype(str)

    main_orders = set(df_main['生产单号'].tolist())

    todo_orders_before = len(df_todo)
    df_todo = df_todo[df_todo['生产单号'].isin(main_orders)].reset_index(drop=True)
    todo_orders_after = len(df_todo)
    removed_count = todo_orders_before - todo_orders_after

    print(f"原始未完成工作表单号数量：{todo_orders_before}")
    print(f"过滤后未完成工作表单号数量：{todo_orders_after}")
    print(f"已删除 {removed_count} 个在大货生产计划表中不存在的单号")

    # ============================================================
    # 目的 1: 将未完成工作表中实际已完成的行移到已完成工作表
    # ============================================================
    print("\n=== 处理目的 1: 移动已完成的工作 ===")

    main_zhengli_shiji = df_main[['生产单号', '整理实际']].copy()
    main_zhengli_shiji.columns = ['生产单号', '整理实际_主表']

    df_todo_merged = df_todo.merge(main_zhengli_shiji, on='生产单号', how='left')

    completed_mask = df_todo_merged['整理实际_主表'].notna()
    completed_rows = df_todo_merged[completed_mask].copy()
    remaining_rows = df_todo_merged[~completed_mask].copy()

    print(f"发现 {len(completed_rows)} 行已完成的工作")
    print(f"剩余 {len(remaining_rows)} 行未完成的工作")

    if len(completed_rows) > 0:
        completed_rows['整理实际'] = completed_rows['整理实际_主表']

        done_cols_without_date = [col for col in df_done.columns if col != '迁移时间']
        move_data = pd.DataFrame(columns=df_done.columns)
        move_data['迁移时间'] = [today] * len(completed_rows)

        for col in done_cols_without_date:
            if col in completed_rows.columns:
                move_data[col] = completed_rows[col].values
            else:
                move_data[col] = None

        df_done_updated = pd.concat([df_done, move_data], ignore_index=True)
        print(f"已将 {len(completed_rows)} 行数据移动到已完成工作表")
    else:
        df_done_updated = df_done.copy()
        print("没有需要移动的数据")

    if '整理实际_主表' in remaining_rows.columns:
        remaining_rows = remaining_rows.drop(columns=['整理实际_主表'])

    # ============================================================
    # 目的 2: 将大货生产计划表中的新单号添加到未完成工作表
    # ============================================================
    print("\n=== 处理目的 2: 添加新单号 ===")

    existing_orders = set()
    existing_orders.update(remaining_rows['生产单号'].astype(str).tolist())
    existing_orders.update(df_done_updated['生产单号'].astype(str).tolist())

    print(f"现有单号数量：{len(existing_orders)}")

    new_orders_mask = ~df_main['生产单号'].astype(str).isin(existing_orders)
    new_orders = df_main[new_orders_mask].copy()

    if '生产车间' in new_orders.columns:
        new_orders = new_orders[new_orders['生产车间'] != '泰州七车间']

    print(f"发现 {len(new_orders)} 个新单号（已排除泰州七车间）")

    if len(new_orders) > 0:
        todo_cols = remaining_rows.columns.tolist()
        new_data = pd.DataFrame()

        for col in todo_cols:
            if col in new_orders.columns:
                new_data[col] = new_orders[col].values
            else:
                new_data[col] = None

        new_data = new_data.dropna(how='all', axis=1)
        df_todo_updated = pd.concat([remaining_rows, new_data], ignore_index=True, sort=False)
        print(f"已将 {len(new_orders)} 个新单号添加到未完成工作表")
    else:
        df_todo_updated = remaining_rows.copy()
        print("没有新单号需要添加")

    # ============================================================
    # 保存结果
    # ============================================================
    print("\n=== 保存结果 ===")

    xls_source = pd.ExcelFile(schedule_backup)
    all_sheets = {}
    for sheet_name in xls_source.sheet_names:
        if sheet_name not in ['未完成', '已完成']:
            all_sheets[sheet_name] = pd.read_excel(schedule_backup, sheet_name=sheet_name)

    output_file = schedule_file.replace('.xlsx', '_temp.xlsx')
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        df_todo_updated.to_excel(writer, sheet_name='未完成', index=False)
        df_done_updated.to_excel(writer, sheet_name='已完成', index=False)

        for sheet_name, df in all_sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    print("正在添加表格边框线...")
    from openpyxl.styles import Border, Side

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    wb = load_workbook(output_file)

    # 使用范围操作一次性设置边框，避免逐个单元格的低效循环
    for sheet_name in ['未完成', '已完成']:
        ws = wb[sheet_name]
        if ws.max_row > 0 and ws.max_column > 0:
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.border = thin_border

        # 单独处理日期列格式
        if sheet_name == '已完成':
            for col_idx, cell in enumerate(ws[1], start=1):
                if cell.value in ['迁移时间', '预排整烫时间']:
                    for row in range(2, ws.max_row + 1):
                        ws.cell(row=row, column=col_idx).number_format = 'yyyy-mm-dd'

    wb.save(output_file)
    print("边框线添加完成，日期格式已设置")

    shutil.move(output_file, schedule_file)

    print("处理完成！结果已保存到后整理计划表.xlsx")
    print(f"\n统计信息:")
    print(f"  - 未完成工作表：{len(df_todo_updated)} 行")
    print(f"  - 已完成工作表：{len(df_done_updated)} 行")
    print(f"  - 移动的行数：{len(completed_rows)}")
    print(f"  - 新增的单号数：{len(new_orders)}")


if __name__ == "__main__":
    create_gui()
