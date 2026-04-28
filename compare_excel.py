#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 表格比较脚本
只保留有差异的数据，物资编码设为文本格式
支持重复物资编码的比较（按出现次数计算差异）
"""

import pandas as pd
import glob
import os
from collections import Counter, defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def compare_excel_files(base_file: str, compare_file: str, output_file: str, key_column: str = None):
    """
    以旧表为基准，比较新表，只输出有差异的数据
    支持重复物资编码：按出现次数计算差异
    """
    # 读取数据
    print(f"正在读取基准文件（旧表）：{base_file}")
    df_base = pd.read_excel(base_file)

    print(f"正在读取比较文件（新表）：{compare_file}")
    df_compare = pd.read_excel(compare_file)

    # 确定关键列
    if key_column is None:
        key_column = df_base.columns[0]

    key_col_idx = list(df_compare.columns).index(key_column)

    print(f"使用关键列进行匹配：{key_column}")

    # 将旧表转换为列表字典（支持重复 key），保留原始行号
    base_list = []
    for idx, row in df_base.iterrows():
        row_dict = row.to_dict()
        row_dict['__原行号__'] = idx + 2  # Excel 行号（从 2 开始，因为第 1 行是表头）
        base_list.append(row_dict)

    # 将新表转换为列表字典（支持重复 key），保留原始行号
    compare_list = []
    for idx, row in df_compare.iterrows():
        row_dict = row.to_dict()
        row_dict['__原行号__'] = idx + 2  # Excel 行号（从 2 开始，因为第 1 行是表头）
        compare_list.append(row_dict)

    # 统计每个物资编码的出现次数
    base_key_counts = Counter(str(row[key_column]) for row in base_list)
    compare_key_counts = Counter(str(row[key_column]) for row in compare_list)

    all_keys = set(base_key_counts.keys()) | set(compare_key_counts.keys())

    # 找出共同行中数据不同的（只比较两边都有的部分，按最小出现次数配对）
    modified_keys = set()
    modify_details = {}  # key -> [(diffs, base_row_num, compare_row_num), ...]

    common_keys = set(base_key_counts.keys()) & set(compare_key_counts.keys())

    for key in common_keys:
        # 获取该 key 在两边的所有行
        base_rows = [row for row in base_list if str(row[key_column]) == key]
        compare_rows = [row for row in compare_list if str(row[key_column]) == key]

        # 按最小出现次数配对比较
        min_count = min(len(base_rows), len(compare_rows))

        for i in range(min_count):
            base_row = base_rows[i]
            compare_row = compare_rows[i]

            diffs = []
            # 排除关键列本身和序号列，只比较实际数据
            for col in df_base.columns:
                if col == key_column:  # 跳过关键列本身
                    continue
                if col == '序号':  # 跳过序号列
                    continue
                val_base = base_row.get(col)
                val_compare = compare_row.get(col)

                if pd.isna(val_base) and pd.isna(val_compare):
                    continue

                base_str = str(val_base) if pd.notna(val_base) else "空"
                compare_str = str(val_compare) if pd.notna(val_compare) else "空"

                if base_str != compare_str:
                    diffs.append(f"{col}: {base_str}→{compare_str}")

            if diffs:
                modified_keys.add(key)
                if key not in modify_details:
                    modify_details[key] = []
                modify_details[key].append({
                    'diffs': diffs,
                    'base_row_num': base_row.get('__原行号__', ''),
                    'compare_row_num': compare_row.get('__原行号__', '')
                })

    # 创建结果列表（只保留有差异的）
    result_rows = []
    columns = list(df_compare.columns) + ['差异说明', '原行号']

    def format_key(val):
        """格式化物资编码：转为文本，去掉.0 后缀"""
        if pd.isna(val):
            return ""
        key_str = str(val)
        if '.0' in key_str:
            key_str = key_str.replace('.0', '')
        return f"'{key_str}"

    # 统计信息
    modified_count = 0
    new_only_count = 0
    old_only_count = 0

    # 1. 新表中与旧表数据不同的行（配对比较后有差异的）
    for key in sorted(modified_keys):
        # 获取新表中该 key 的所有行
        compare_rows = [row for row in compare_list if str(row[key_column]) == key]
        for detail in modify_details[key]:
            row_data = compare_rows[0]  # 取第一行作为代表
            new_row = []
            for col in df_compare.columns:
                val = row_data.get(col, None)
                if col == key_column:
                    val = format_key(val)
                new_row.append(val)
            details = "; ".join(detail['diffs'][:5])
            if len(detail['diffs']) > 5:
                details += f"...等共{len(detail['diffs'])}处差异"
            new_row.append(f"【数据不同】{details}")
            new_row.append(f"旧表:{detail['base_row_num']} 新表:{detail['compare_row_num']}")
            result_rows.append(new_row)
            modified_count += 1

    # 2. 按数量差异计算：新表比旧表多的部分（新表独有）
    for key in all_keys:
        base_count = base_key_counts.get(key, 0)
        compare_count = compare_key_counts.get(key, 0)
        
        if compare_count > base_count:
            # 新表比旧表多出的行数
            extra_count = compare_count - base_count
            # 获取新表中该 key 的所有行，取多出的部分
            compare_rows = [row for row in compare_list if str(row[key_column]) == key]
            for row_data in compare_rows[base_count:]:  # 跳过与旧表配对的部分
                new_row = []
                for col in df_compare.columns:
                    val = row_data.get(col, None)
                    if col == key_column:
                        val = format_key(val)
                    new_row.append(val)
                new_row.append("【新表多出】旧表数量不足，新表多出此记录")
                new_row.append(f"新表:{row_data.get('__原行号__','')}")
                result_rows.append(new_row)
                new_only_count += 1

    # 3. 按数量差异计算：旧表比新表多的部分（旧表独有）- 追加在末尾
    for key in all_keys:
        base_count = base_key_counts.get(key, 0)
        compare_count = compare_key_counts.get(key, 0)
        
        if base_count > compare_count:
            # 旧表比新表多出的行数
            extra_count = base_count - compare_count
            # 获取旧表中该 key 的所有行，取多出的部分
            base_rows = [row for row in base_list if str(row[key_column]) == key]
            for row_data in base_rows[compare_count:]:  # 跳过与新表配对的部分
                new_row = []
                for col in df_compare.columns:
                    val = row_data.get(col, None)
                    if col == key_column:
                        val = format_key(val)
                    new_row.append(val)
                new_row.append("【旧表多出】新表数量不足，旧表多出此记录")
                new_row.append(f"旧表:{row_data.get('__原行号__','')}")
                result_rows.append(new_row)
                old_only_count += 1

    # 创建 Workbook 并写入数据
    wb = Workbook()
    ws = wb.active
    ws.title = "差异对比"

    # 设置表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")

    for col, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 写入数据
    for row_idx, row_data in enumerate(result_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            # 物资编码列设为文本格式
            if col_idx == key_col_idx + 1:
                cell.number_format = '@'

    # 设置差异说明列的样式
    different_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 黄色 - 数据不同
    unique_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # 绿色 - 新表多出
    missing_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # 红色 - 旧表多出

    diff_col = len(columns) - 1  # 差异说明列
    row_col = len(columns)       # 原行号列

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=diff_col)
        value = cell.value if cell.value else ""

        if "【数据不同】" in value:
            for col in range(1, len(ws[row]) + 1):
                ws.cell(row=row, column=col).fill = different_fill
        elif "【新表多出】" in value:
            for col in range(1, len(ws[row]) + 1):
                ws.cell(row=row, column=col).fill = unique_fill
        elif "【旧表多出】" in value:
            for col in range(1, len(ws[row]) + 1):
                ws.cell(row=row, column=col).fill = missing_fill

    # 设置列宽
    for i in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.column_dimensions[get_column_letter(diff_col)].width = 50

    # 冻结首行
    ws.freeze_panes = "A2"

    wb.save(output_file)

    # 打印统计
    print(f"\n{'='*60}")
    print(f"比较完成！")
    print(f"{'='*60}")
    print(f"旧表总行数：{len(df_base)}")
    print(f"新表总行数：{len(df_compare)}")
    print(f"-------------------------------------------")
    print(f"有差异的记录总数：{len(result_rows)}")
    print(f"  - 数据有不同：{modified_count} 行")
    print(f"  - 新表多出（数量超出旧表）：{new_only_count} 行")
    print(f"  - 旧表多出（数量超出新表）：{old_only_count} 行")
    print(f"{'='*60}")
    print(f"结果已保存至：{output_file}")
    print(f"物资编码列已设置为文本格式")
    print(f"{'='*60}")


def main():
    # 使用脚本所在目录作为工作目录（兼容 Windows/Linux）
    script_dir = os.path.dirname(os.path.abspath(__file__))
    os.chdir(script_dir)

    # 明确指定旧表和新表文件名
    base_file = '旧表.xlsx'
    compare_file = '新表.xlsx'

    # 检查文件是否存在
    if not os.path.exists(base_file):
        print(f"错误：找不到文件 {base_file}")
        return
    if not os.path.exists(compare_file):
        print(f"错误：找不到文件 {compare_file}")
        return

    # 读取列名供用户选择
    df_temp = pd.read_excel(base_file)
    columns = list(df_temp.columns)

    print("\n========== Excel 表格比较工具 ==========")
    print(f"旧表：{base_file}")
    print(f"新表：{compare_file}")
    print(f"\n可用的列名：{columns}")
    print()

    # 用户输入关键列
    while True:
        key_column = input("请输入用于匹配的关键列名（直接回车使用第一列）：").strip()
        if not key_column:
            key_column = columns[0]
            print(f"已选择第一列：{key_column}")
            break
        if key_column in columns:
            print(f"已选择：{key_column}")
            break
        print(f"错误：列名 '{key_column}' 不存在，请重新输入！")

    compare_excel_files(base_file, compare_file, '差异对比结果.xlsx', key_column=key_column)


if __name__ == "__main__":
    main()
